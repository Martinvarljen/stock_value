"""
LEGACY RESEARCH — not the production paper agent (see text/STRATEGY_MAP.md).
For research_ls use portfolio/backtest.py and portfolio/daily_run.py instead.

strategy_backtest.py  —  Historical strategy backtesting engine

Tests whether the stock_analyzer classification signals predicted future returns.

Approach:
  1. For each ticker, pull all available financial data from yfinance (5Y history).
  2. For checkpoint dates (quarter-ends by default, or month-ends with ``--checkpoint-freq M``):
     - Filter financials to only what was publicly available at that date
       (fiscal year end + 90 day reporting lag).
     - Get the historical closing price at that date.
     - Reconstruct the data dict and run the full analysis pipeline.
     - Record the classification.
  3. Measure forward returns (3M, 6M, 12M) from each checkpoint.
  4. Optionally measure the same forward returns on a benchmark (default SPY)
     and store excess return (stock minus benchmark) per signal.
  5. Aggregate: average return by classification tier, hit rate, etc.

Usage:
    python backtesting/strategy_backtest.py              # default: yearly top-100 universe (needs cache)
    python backtesting/strategy_backtest.py --classic-tickers   # old 16-name demo list
    python backtesting/strategy_backtest.py AAPL MSFT    # explicit tickers (no yearly universe)
    python strategy_backtest.py --lookback 3          # 3 years lookback
    python strategy_backtest.py --tickers-file tickers.txt
    python strategy_backtest.py --benchmark SPY       # default; use --benchmark none to disable
    python backtesting/strategy_backtest.py --no-valuation   # dcf mode without DCF (classification_engine only)

Dynamic entry/exit (default HTML when benchmark is on — same run as above):
    Writes ``strategy_dynamic_vs_spy_YYYYMMDD.html`` (animated monthly strategy vs SPY buy-and-hold).
    Optional: ``--legacy-html`` for the older diagnostic chart (scatter / quarterly / weighted basket).

Yearly top-100 universe (lagged one calendar year, no look-ahead), checkpoints from 2023 through today vs SPY by default:
    python backtesting/build_yearly_top100_universe.py --for-checkpoints-from-year 2023
    python backtesting/strategy_backtest.py --yearly-top100
    python backtesting/strategy_backtest.py --yearly-top100 --checkpoint-freq M
    # Month-end checkpoints (~3x more evaluation dates than quarters).
    python backtesting/strategy_backtest.py --yearly-top100 --strategy ml
    # ML projection strategy: technicals + Dolt-trained LightGBM (projection_engine, no DCF).
    python backtesting/strategy_backtest.py --yearly-top100 --signal-tech-ai
    # Same as --strategy ml (legacy flag).

    Each checkpoint in year C only evaluates names in the top-100 list for year C-1
    (built from prior-year dollar volume among current S&P 500 — see yearly_top100_universe.py).
    Optional: --auto-build-universe  to build missing year files (slow, many API calls).
"""

import sys
import os
import math
import logging
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from pathlib import Path

# Add repo root (for backtesting.*) and stock_analyzer to import path
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "stock_analyzer"))
if str(_ROOT / "projection") not in sys.path:
    sys.path.insert(0, str(_ROOT / "projection"))

from quality_engine import analyze_quality

from backtesting.overlap_stats import effective_n_for_signals, overlap_inflation_factor
from backtesting.yearly_top100_universe import (
    UniverseSource,
    default_universe_cache_dir,
    load_universe_map_for_lag_years,
    normalize_universe_source,
)
from backtesting.strategy_modes import (
    MODE_DCF,
    is_ml_strategy,
    normalize_signal_mode,
    strategy_display_name,
)
logger = logging.getLogger(__name__)

from backtesting.ml_quant import (
    aggregate_quintile_forward_returns,
    ml_score_from_signal,
    print_quintile_table,
)
from financial_strength import analyze_financials
from valuation_engine import analyze_valuation
from growth_engine import analyze_growth
from risk_engine import analyze_risk
from red_flags import analyze_red_flags
from classification_engine import classify_stock
from sector_engine import apply_sector_context
from momentum_engine import analyze_momentum
from technical_extended import analyze_extended_technicals
from projection_engine import generate_projections
from utils import _pct

# ── console encoding fix (Windows) ─────────────────────────────────────────────

def _configure_console():
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass

_configure_console()


# ── constants ──────────────────────────────────────────────────────────────────

REPORTING_LAG_DAYS = 90       # assume financials available 90 days after fiscal year end
MARGIN_OF_SAFETY   = 0.30
DEFAULT_LOOKBACK   = 2        # years
# When using yearly top-100 universe, only evaluate checkpoints from this calendar year onward (S&P proxy universe is year-lagged).
DEFAULT_TOP100_CHECKPOINT_MIN_YEAR = 2023
CHECKPOINT_FREQ    = "Q"      # default: quarter-end; use "M" for month-end (see --checkpoint-freq)
FORWARD_PERIODS    = [3, 6, 12]  # months
# Default simulated hold for basket equity curve + dynamic backtest (quarterly CP → 3M non-overlap).
DEFAULT_HOLD_MONTHS = 3
# Clip each signal's forward return when computing tier *averages/medians* (Yahoo outliers).
TIER_STATS_WINSOR = 0.5  # ±50%

TIER_ORDER = ["STRONG BUY", "BUY", "WATCHLIST", "HOLD", "AVOID", "STRONG AVOID"]
BUY_TIERS = frozenset({"STRONG BUY", "BUY"})


def hold_days(months: int) -> int:
    return int(months * 30.44)


def _fwd_keys(months: int) -> tuple[str, str]:
    return f"fwd_{months}m", f"spy_fwd_{months}m"


def weighted_basket_returns_at_date(
    signals_same_date: list[dict],
    *,
    weight_mode: str = "tier",
    hold_months: int = DEFAULT_HOLD_MONTHS,
) -> tuple[float | None, float | None, int]:
    """
    One-day basket among ``signals_same_date``: return (stock_fwd, spy_fwd, n_names) or (None,None,0).

    Uses ``fwd_{hold_months}m`` / ``spy_fwd_{hold_months}m`` (must match simulated hold).
  ``weight_mode``: ``tier`` (STRONG BUY weight 2, BUY weight 1) or ``equal``.
    """
    fk, sk = _fwd_keys(hold_months)
    eligible = [
        s
        for s in signals_same_date
        if s.get("classification") in BUY_TIERS
        and s.get(fk) is not None
        and s.get(sk) is not None
    ]
    if not eligible:
        return None, None, 0
    if weight_mode == "equal":
        w = [1.0 / len(eligible)] * len(eligible)
    else:
        raw = [2.0 if s["classification"] == "STRONG BUY" else 1.0 for s in eligible]
        tw = sum(raw)
        w = [x / tw for x in raw]
    r_s = sum(wi * float(s[fk]) for wi, s in zip(w, eligible))
    r_b = sum(wi * float(s[sk]) for wi, s in zip(w, eligible))
    return r_s, r_b, len(eligible)


def sequential_weighted_equity_curve(
    signals: list[dict],
    *,
    weight_mode: str = "tier",
    hold_months: int = DEFAULT_HOLD_MONTHS,
) -> pd.DataFrame:
    """
    Non-overlapping holds of ``hold_months``: at each entry, invest in a weighted basket of
    all BUY-tier signals **on that checkpoint date**; realize ``fwd_{hold_months}m`` after
    ``hold_days(hold_months)``; next entry is the earliest later checkpoint on or after exit.

    Returns columns: entry_date, exit_date, n_names, ret_stock, ret_spy, equity_stock, equity_spy.
    """
    fk, sk = _fwd_keys(hold_months)
    hd = hold_days(hold_months)
    by_date = {}
    for s in signals:
        if s.get(fk) is None or s.get(sk) is None:
            continue
        d = s["date"]
        d = pd.Timestamp(d).to_pydatetime()
        if d.tzinfo is not None:
            d = d.replace(tzinfo=None)
        by_date.setdefault(d, []).append(s)

    dates = sorted(by_date.keys())
    if not dates:
        return pd.DataFrame(
            columns=["entry_date", "exit_date", "n_names", "ret_stock", "ret_spy", "equity_stock", "equity_spy"]
        )

    rows: list[dict] = []
    eq_s = 1.0
    eq_b = 1.0
    exit_after: datetime | None = None

    for d in dates:
        if exit_after is not None and d < exit_after:
            continue
        r_s, r_b, n = weighted_basket_returns_at_date(
            by_date[d], weight_mode=weight_mode, hold_months=hold_months
        )
        if r_s is None or r_b is None or n == 0:
            continue
        ex = d + timedelta(days=hd)
        eq_s *= 1.0 + r_s
        eq_b *= 1.0 + r_b
        rows.append(
            {
                "entry_date": d,
                "exit_date": ex,
                "n_names": n,
                "ret_stock": r_s,
                "ret_spy": r_b,
                "equity_stock": eq_s,
                "equity_spy": eq_b,
            }
        )
        exit_after = ex

    return pd.DataFrame(rows)


# ── data collection ───────────────────────────────────────────────────────────

def _safe(value, default=None):
    try:
        if value is None:
            return default
        if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
            return default
        return value
    except Exception:
        return default


def _series_to_list(series, n=5):
    if series is None or series.empty:
        return [None] * n
    series = series.dropna().sort_index()
    values = [_safe(v) for v in series.values]
    if len(values) >= n:
        return values[-n:]
    return [None] * (n - len(values)) + values


def _cagr(start, end, years):
    try:
        if start is None or end is None or years <= 0 or start <= 0 or end <= 0:
            return None
        r = (end / start) ** (1 / years) - 1
        if not math.isfinite(r):
            return None
        return r
    except Exception:
        return None


def _pct_change_list(lst):
    changes = []
    for i in range(1, len(lst)):
        a, b = lst[i - 1], lst[i]
        if a is None or b is None or a == 0:
            changes.append(None)
        else:
            changes.append((b - a) / abs(a))
    return changes


def _list_cagr(lst, require_positive_start=False):
    indexed = [(i, v) for i, v in enumerate(lst) if v is not None]
    if len(indexed) < 2:
        return None
    _, v_start = indexed[0]
    _, v_end = indexed[-1]
    years = indexed[-1][0] - indexed[0][0]
    if years <= 0 or (require_positive_start and v_start <= 0):
        return None
    return _cagr(v_start, v_end, years)


def _row(df, *keys):
    if df is None or df.empty:
        return pd.Series(dtype=float)
    for k in keys:
        if k in df.index:
            return df.loc[k]
    return pd.Series(dtype=float)


def _last(lst):
    for v in reversed(lst):
        if v is not None:
            return v
    return None


# ── point-in-time data reconstruction ─────────────────────────────────────────

def _filter_before(df, cutoff_date):
    """Filter a financials DataFrame to only columns (fiscal periods) available before cutoff."""
    if df is None or df.empty:
        return None
    cutoff = pd.Timestamp(cutoff_date)
    valid_cols = []
    for c in df.columns:
        col_ts = pd.Timestamp(c)
        if col_ts.tz is not None:
            col_ts = col_ts.tz_localize(None)
        if col_ts + pd.Timedelta(days=REPORTING_LAG_DAYS) <= cutoff:
            valid_cols.append(c)
    if not valid_cols:
        return None
    return df[valid_cols].sort_index(axis=1)


def _get_price_at(price_history: pd.DataFrame, target_date: datetime) -> float | None:
    """Get the closing price closest to (but not after) target_date."""
    if price_history is None or price_history.empty:
        return None
    target = pd.Timestamp(target_date)
    idx = price_history.index.tz_localize(None) if price_history.index.tz else price_history.index
    valid = price_history[idx <= target]
    if valid.empty:
        return None
    return float(valid["Close"].iloc[-1])


def _get_forward_price(price_history: pd.DataFrame, from_date: datetime, months: int) -> float | None:
    """Closing price at the first bar on/after the target date (~months forward).

    Returns None when history does not extend far enough (no fallback to latest bar).
    """
    if price_history is None or price_history.empty:
        return None
    idx = price_history.index.tz_localize(None) if price_history.index.tz else price_history.index
    target = pd.Timestamp(from_date + timedelta(days=months * 30.44))
    if getattr(target, "tz", None) is not None:
        target = target.tz_localize(None)
    future = price_history[idx >= target]
    if future.empty:
        return None
    first_idx = pd.Timestamp(future.index[0])
    if getattr(first_idx, "tz", None) is not None:
        first_idx = first_idx.tz_localize(None)
    if abs((first_idx - target).days) > 15:
        return None
    return float(future["Close"].iloc[0])


def _load_tickers_from_file(path: str) -> list[str]:
    """One ticker per line; blank lines and # comments ignored."""
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"Tickers file not found: {path}")
    out: list[str] = []
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        sym = line.split()[0].strip().upper()
        if sym:
            out.append(sym)
    return out


def _load_benchmark_history(benchmark: str) -> pd.DataFrame | None:
    """Daily history for benchmark (e.g. SPY) for aligned forward returns."""
    if not benchmark or benchmark.lower() in ("none", "off", "false", "-"):
        return None
    try:
        h = yf.Ticker(benchmark).history(period="max", interval="1d")
        if h is None or h.empty or "Close" not in h.columns:
            return None
        h = h.copy()
        if h.index.tz is not None:
            h.index = h.index.tz_localize(None)
        return h
    except Exception:
        return None


def collect_raw_yfinance(ticker: str) -> dict:
    """Pull all raw yfinance data once per ticker (financials + full price history)."""
    tk = yf.Ticker(ticker)
    info = tk.info or {}

    try:
        inc = tk.financials
        inc_a = inc.sort_index(axis=1) if inc is not None and not inc.empty else None
    except Exception:
        inc_a = None

    try:
        bal = tk.balance_sheet
        bal_a = bal.sort_index(axis=1) if bal is not None and not bal.empty else None
    except Exception:
        bal_a = None

    try:
        cf = tk.cashflow
        cf_a = cf.sort_index(axis=1) if cf is not None and not cf.empty else None
    except Exception:
        cf_a = None

    # Pull full price history (daily) for the lookback + forward period
    try:
        hist = tk.history(period="max", interval="1d")
    except Exception:
        hist = pd.DataFrame()

    return {
        "ticker": ticker,
        "info": info,
        "income_statement": inc_a,
        "balance_sheet": bal_a,
        "cash_flow": cf_a,
        "price_history": hist,
    }


def reconstruct_price_only_at(raw: dict, as_of_date: datetime) -> dict | None:
    """
    Minimal point-in-time record for ML scoring when yfinance fundamentals
    are too sparse (common before ~2020 on ``income_statement``).
    """
    ticker = raw.get("ticker", "")
    price = _get_price_at(raw.get("price_history"), as_of_date)
    if price is None or price <= 0:
        return None

    info = raw.get("info") or {}
    result = {
        "ticker": ticker,
        "error": None,
        "current_price": price,
        "beta": _safe(info.get("beta")),
        "feature_as_of": as_of_date,
        "checkpoint_date": as_of_date,
    }
    enrich_point_in_time_technicals(result, raw, as_of_date)
    # ml_model.features.MIN_OHLCV_BARS
    if len(result.get("close_1y") or []) < 220:
        return None
    return result


def reconstruct_data_at(raw: dict, as_of_date: datetime) -> dict | None:
    """
    Reconstruct a data dict as it would have appeared on `as_of_date`.
    Only uses financials published before that date and the price on that date.
    Returns None if insufficient data.
    """
    info = raw["info"]
    ticker = raw["ticker"]

    # Filter financials to point-in-time
    inc_a = _filter_before(raw["income_statement"], as_of_date)
    bal_a = _filter_before(raw["balance_sheet"], as_of_date)
    cf_a = _filter_before(raw["cash_flow"], as_of_date)

    if inc_a is None or inc_a.shape[1] < 2:
        return None

    # Get price at checkpoint date
    price = _get_price_at(raw["price_history"], as_of_date)
    if price is None or price <= 0:
        return None

    # Handle GBp → GBP
    currency = _safe(info.get("currency"))
    if currency == "GBp":
        price = price / 100
        currency = "GBP"

    result = {
        "ticker": ticker,
        "error": None,
        "company_name": _safe(info.get("longName") or info.get("shortName")),
        "quote_type": _safe(info.get("quoteType")),
        "sector": _safe(info.get("sector")),
        "industry": _safe(info.get("industry")),
        "exchange": _safe(info.get("exchange")),
        "currency": currency,
        "current_price": price,
        "eur_rate": None,
        "price_eur": None,
    }

    # Income statement rows
    rev_s = _row(inc_a, "Total Revenue", "Revenue")
    gp_s = _row(inc_a, "Gross Profit")
    ebit_s = _row(inc_a, "EBIT", "Operating Income")
    ni_s = _row(inc_a, "Net Income")
    eps_s = _row(inc_a, "Basic EPS", "Diluted EPS")
    int_s = _row(inc_a, "Interest Expense")
    dep_s = _row(inc_a, "Reconciled Depreciation", "Depreciation And Amortization")
    tax_s = _row(inc_a, "Tax Provision", "Income Tax Expense")
    pbt_s = _row(inc_a, "Pretax Income")

    # Balance sheet rows
    assets_s = _row(bal_a, "Total Assets")
    liab_s = _row(bal_a, "Total Liabilities Net Minority Interest", "Total Liabilities")
    ltd_s = _row(bal_a, "Long Term Debt")
    std_s = _row(bal_a, "Current Debt", "Short Term Debt", "Current Portion Of Long Term Debt And Capital Lease Obligation")
    cash_s = _row(bal_a, "Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments")
    equity_s = _row(bal_a, "Stockholders Equity", "Total Stockholders Equity", "Common Stock Equity")
    shares_s = _row(bal_a, "Share Issued", "Ordinary Shares Number")
    cur_assets = _row(bal_a, "Current Assets")
    cur_liab = _row(bal_a, "Current Liabilities")

    # Cash flow rows
    ocf_s = _row(cf_a, "Operating Cash Flow", "Cash Flow From Continuing Operating Activities")
    capex_s = _row(cf_a, "Capital Expenditure")
    div_s = _row(cf_a, "Common Stock Dividend Paid", "Cash Dividends Paid")
    sbc_s = _row(cf_a, "Stock Based Compensation")

    # Store 5-year annual lists
    result["revenue_5y"] = _series_to_list(rev_s)
    result["gross_profit_5y"] = _series_to_list(gp_s)
    result["ebit_5y"] = _series_to_list(ebit_s)
    result["net_income_5y"] = _series_to_list(ni_s)
    result["eps_5y"] = _series_to_list(eps_s)
    result["interest_expense_5y"] = _series_to_list(int_s)
    result["depreciation_5y"] = _series_to_list(dep_s)
    result["tax_provision_5y"] = _series_to_list(tax_s)
    result["pretax_income_5y"] = _series_to_list(pbt_s)

    result["total_assets_5y"] = _series_to_list(assets_s)
    result["total_liabilities_5y"] = _series_to_list(liab_s)
    result["total_debt_5y"] = _series_to_list(
        (ltd_s.fillna(0) + std_s.fillna(0)) if (not ltd_s.empty or not std_s.empty) else pd.Series(dtype=float)
    )
    result["cash_5y"] = _series_to_list(cash_s)
    result["equity_5y"] = _series_to_list(equity_s)
    result["shares_5y"] = _series_to_list(shares_s)
    # Point-in-time shares from balance sheet; yfinance.info is today's snapshot only.
    shares_outstanding = _last(result["shares_5y"]) or _safe(info.get("sharesOutstanding"))
    result["shares_outstanding"] = shares_outstanding
    result["market_cap"] = (shares_outstanding * price) if shares_outstanding else None
    result["beta"] = _safe(info.get("beta"))  # still today's beta from info (no historical field)

    result["current_assets_5y"] = _series_to_list(cur_assets)
    result["current_liab_5y"] = _series_to_list(cur_liab)
    result["ocf_5y"] = _series_to_list(ocf_s)
    result["capex_5y"] = _series_to_list(capex_s)
    result["dividends_5y"] = _series_to_list(div_s)
    result["sbc_5y"] = _series_to_list(sbc_s)

    # FCF
    fcf_list = []
    for o, c in zip(result["ocf_5y"], result["capex_5y"]):
        if o is None:
            fcf_list.append(None)
        elif c is None:
            fcf_list.append(o)
        else:
            fcf_list.append(o - abs(c))
    result["fcf_5y"] = fcf_list

    # Latest-year snapshots
    revenue = _last(result["revenue_5y"])
    gross_profit = _last(result["gross_profit_5y"])
    ebit = _last(result["ebit_5y"])
    net_income = _last(result["net_income_5y"])
    ocf = _last(result["ocf_5y"])
    capex = _last(result["capex_5y"])
    fcf = _last(result["fcf_5y"])
    total_debt = _last(result["total_debt_5y"])
    cash = _last(result["cash_5y"])
    equity = _last(result["equity_5y"])
    interest_exp = _last(result["interest_expense_5y"])
    dep = _last(result["depreciation_5y"])
    cur_a = _last(result["current_assets_5y"])
    cur_l = _last(result["current_liab_5y"])
    tax_prov = _last(result["tax_provision_5y"])
    pretax = _last(result["pretax_income_5y"])

    result.update({
        "revenue": revenue, "gross_profit": gross_profit,
        "ebit": ebit, "net_income": net_income,
        "ocf": ocf, "capex": capex, "fcf": fcf,
        "total_debt": total_debt, "cash": cash,
        "equity": equity, "interest_expense": interest_exp,
        "depreciation": dep,
    })

    # Derived ratios
    mktcap = result["market_cap"]
    ev = (mktcap + (total_debt or 0) - (cash or 0)) if mktcap else None
    result["enterprise_value"] = ev

    result["gross_margin"] = _safe(gross_profit / revenue) if (revenue not in (None, 0) and gross_profit is not None) else None
    result["operating_margin"] = _safe(ebit / revenue) if (revenue not in (None, 0) and ebit is not None) else None
    result["net_margin"] = _safe(net_income / revenue) if (revenue not in (None, 0) and net_income is not None) else None

    ebitda = None
    if ebit is not None and dep is not None:
        ebitda = ebit + dep
    elif ebit is not None:
        ebitda = ebit
    result["ebitda"] = ebitda

    result["roe"] = _safe(net_income / equity) if (net_income is not None and equity not in (None, 0)) else None
    assets_last = _last(result["total_assets_5y"])
    result["roa"] = _safe(net_income / assets_last) if (net_income is not None and assets_last not in (None, 0)) else None

    eff_tax = None
    if tax_prov is not None and pretax is not None and pretax != 0:
        eff_tax = max(0.0, min(0.5, tax_prov / pretax))
    result["effective_tax_rate"] = _safe(eff_tax, 0.22)

    net_debt = (total_debt - cash) if (total_debt is not None and cash is not None) else total_debt
    result["net_debt"] = net_debt
    invested_capital = (equity + (net_debt or 0)) if equity else None
    if ebit is not None and invested_capital and invested_capital != 0:
        t = result["effective_tax_rate"] or 0.22
        result["roic"] = _safe(ebit * (1 - t) / invested_capital)
    else:
        result["roic"] = None

    result["debt_equity"] = _safe(total_debt / equity) if (total_debt is not None and equity not in (None, 0)) else None
    result["net_debt_ebitda"] = _safe(net_debt / ebitda) if (net_debt is not None and ebitda not in (None, 0)) else None
    result["interest_coverage"] = _safe(abs(ebit) / abs(interest_exp)) if (ebit is not None and interest_exp not in (None, 0)) else None
    result["current_ratio"] = _safe(cur_a / cur_l) if (cur_a is not None and cur_l not in (None, 0)) else None
    result["fcf_yield"] = _safe(fcf / mktcap) if (fcf is not None and mktcap not in (None, 0)) else None

    result["ev_ebit"] = _safe(ev / ebit) if (ev is not None and ebit not in (None, 0)) else None
    result["ev_ebitda"] = _safe(ev / ebitda) if (ev is not None and ebitda not in (None, 0)) else None
    result["ev_fcf"] = _safe(ev / fcf) if (ev is not None and fcf not in (None, 0)) else None

    # Approximate P/E using point-in-time price and trailing EPS
    trailing_eps = _last(result["eps_5y"])
    result["trailing_eps"] = trailing_eps
    result["trailing_pe"] = _safe(price / trailing_eps) if (trailing_eps and trailing_eps > 0) else None
    result["forward_pe"] = None
    result["pb_ratio"] = _safe(price * (shares_outstanding or 0) / equity) if (equity and equity > 0 and shares_outstanding) else None
    result["dividend_yield"] = _safe(info.get("dividendYield"))
    result["peg_ratio"] = None
    result["ev_revenue"] = _safe(ev / revenue) if (ev is not None and revenue not in (None, 0)) else None

    # Growth CAGRs
    result["revenue_cagr_5y"] = _list_cagr(result["revenue_5y"])
    result["eps_cagr_5y"] = _list_cagr(result["eps_5y"], require_positive_start=True)
    result["fcf_cagr_5y"] = _list_cagr(result["fcf_5y"], require_positive_start=True)
    result["revenue_yoy_changes"] = _pct_change_list(result["revenue_5y"])

    # Shares dilution
    sh_now = shares_outstanding or _last(result["shares_5y"])
    sh_old = next((v for v in result["shares_5y"] if v is not None), None)
    if sh_now and sh_old and sh_old != 0:
        result["shares_change_pct"] = (sh_now - sh_old) / sh_old
    else:
        result["shares_change_pct"] = None

    # Capex ratios
    capex_ratios = []
    for r, c in zip(result["revenue_5y"], result["capex_5y"]):
        if r and c and r != 0:
            capex_ratios.append(abs(c) / r)
    result["capex_pct_revenue"] = float(np.mean(capex_ratios)) if capex_ratios else None

    dep_5y = result.get("depreciation_5y") or []
    net_capex_ratios = []
    for r, c, d in zip(result["revenue_5y"], result["capex_5y"], dep_5y):
        if r and c and r != 0:
            gross_cx = abs(c)
            da = abs(d) if d else 0.0
            net_cx = max(gross_cx - da, 0.0)
            net_capex_ratios.append(net_cx / r)
    result["net_capex_pct_revenue"] = float(np.mean(net_capex_ratios)) if net_capex_ratios else None

    # Price history fields (approximated from what we have up to as_of_date)
    ph = raw["price_history"]
    if ph is not None and not ph.empty:
        cutoff = pd.Timestamp(as_of_date)
        ph_idx = ph.index.tz_localize(None) if ph.index.tz else ph.index
        hist_before = ph[ph_idx <= cutoff]
        if len(hist_before) >= 200:
            result["ma50"] = float(hist_before["Close"].iloc[-50:].mean())
            result["ma200"] = float(hist_before["Close"].iloc[-200:].mean())
        else:
            result["ma50"] = None
            result["ma200"] = None

        # 1Y return
        if len(hist_before) > 252:
            p_1y_ago = float(hist_before["Close"].iloc[-252])
            result["return_1y"] = (price - p_1y_ago) / p_1y_ago if p_1y_ago > 0 else None
        else:
            result["return_1y"] = None
        result["return_3y"] = None
    else:
        result["ma50"] = None
        result["ma200"] = None
        result["return_1y"] = None
        result["return_3y"] = None

    result["rsi14"] = None
    result["close_1y"] = []
    result["close_5y_monthly"] = []
    result["week52_high"] = None
    result["week52_low"] = None
    result["pe_5y_min"] = None
    result["pe_5y_max"] = None
    result["pe_5y_median"] = None
    result["data_quality_score"] = 60  # approximate since we have limited data

    return result


def _rsi14_from_closes(closes: list[float] | np.ndarray) -> float | None:
    ser = pd.Series(closes, dtype=float)
    if len(ser) < 15:
        return None
    delta = ser.diff()
    gain = delta.clip(lower=0.0)
    loss = (-delta).clip(lower=0.0)
    avg_g = gain.ewm(alpha=1.0 / 14.0, adjust=False).mean()
    avg_l = loss.ewm(alpha=1.0 / 14.0, adjust=False).mean()
    rs = avg_g / avg_l.replace(0, np.nan)
    rsi = 100.0 - (100.0 / (1.0 + rs))
    v = float(rsi.iloc[-1])
    return v if math.isfinite(v) else None


def enrich_point_in_time_technicals(data: dict, raw: dict, as_of: datetime) -> None:
    """Fill ``data`` price fields from ``raw['price_history']`` through ``as_of`` (mutates ``data``)."""
    ph = raw.get("price_history")
    if ph is None or getattr(ph, "empty", True) or "Close" not in ph.columns:
        return
    cutoff = pd.Timestamp(as_of)
    sub = ph.sort_index()
    try:
        if getattr(sub.index, "tz", None) is not None:
            sub = sub.tz_convert("UTC").tz_localize(None)
    except Exception:
        sub = ph.sort_index()
    sub = sub.loc[sub.index <= cutoff]
    if sub.empty:
        return
    n = len(sub)
    closes = sub["Close"].astype(float)
    last_n = min(252, n)
    tail = sub.iloc[-last_n:]
    data["close_1y"] = tail["Close"].astype(float).tolist()
    if "High" in tail.columns:
        data["high_1y"] = tail["High"].astype(float).tolist()
    else:
        data["high_1y"] = list(data["close_1y"])
    if "Low" in tail.columns:
        data["low_1y"] = tail["Low"].astype(float).tolist()
    else:
        data["low_1y"] = list(data["close_1y"])
    if "Volume" in tail.columns:
        data["volume_1y"] = tail["Volume"].astype(float).fillna(0.0).tolist()
    else:
        data["volume_1y"] = [1.0] * len(data["close_1y"])

    arr = closes.to_numpy(dtype=float)
    data["ma50"] = float(closes.iloc[-50:].mean()) if n >= 50 else None
    data["ma200"] = float(closes.iloc[-200:].mean()) if n >= 200 else None
    data["rsi14"] = _rsi14_from_closes(arr)

    if n >= 5 and isinstance(sub.index, pd.DatetimeIndex):
        mc = sub["Close"].astype(float).resample("ME").last().dropna()
        data["close_5y_mo"] = mc.tail(60).tolist()
    else:
        data["close_5y_mo"] = []

    span = min(252, n)
    w52 = sub.iloc[-span:]
    if "High" in w52.columns:
        data["week52_high"] = float(w52["High"].max())
        data["week52_low"] = float(w52["Low"].min())
    else:
        data["week52_high"] = float(w52["Close"].max())
        data["week52_low"] = float(w52["Close"].min())

    if n >= 253:
        p0, p1 = float(closes.iloc[-253]), float(closes.iloc[-1])
        if p0 > 0:
            data["return_1y"] = (p1 - p0) / p0


def projection_signal_to_classification(sig: str) -> str:
    return {
        "BULLISH": "STRONG BUY",
        "LEAN_BULLISH": "BUY",
        "NEUTRAL": "HOLD",
        "LEAN_BEARISH": "AVOID",
        "BEARISH": "STRONG AVOID",
    }.get(sig or "", "HOLD")


def infer_classification_tech_ai(
    data: dict,
    raw: dict,
    as_of: datetime,
    *,
    spy_close_series: pd.Series | None = None,
) -> tuple[str | None, dict]:
    """
    Technicals + fundamentals (no DCF) + projection_engine (ML blend when available).
    Maps projection ``signal`` onto the same tier labels used by the DCF classifier.
    """
    meta: dict = {"signal_mode": "tech_ai", "strategy": "ml"}
    data["feature_as_of"] = as_of
    data["checkpoint_date"] = as_of
    if spy_close_series is not None and len(spy_close_series):
        data["spy_close_series"] = spy_close_series
    enrich_point_in_time_technicals(data, raw, as_of)

    sector_result = apply_sector_context(data)
    analyze_quality(data)
    financial_result = analyze_financials(data)
    analyze_growth(data)
    risk_result = analyze_risk(data)
    red_flag_result = analyze_red_flags(data, wacc=None)

    mom = analyze_momentum(data)
    data["momentum_metrics"] = mom["momentum_metrics"]
    data["momentum_flags"] = mom["momentum_flags"]
    data["momentum_trend"] = mom["trend"]
    data["extended_technicals"] = analyze_extended_technicals(data)

    all_critical = (
        (financial_result.get("critical_flags") or []) +
        (risk_result.get("critical_flags") or [])
    )
    record = {
        **data,
        "fair_value_weighted": None,
        "buy_below_price": None,
        "valuation_metrics": {},
        "wacc_data": {},
        "red_flags": red_flag_result["red_flags"],
        "critical_flags": all_critical,
        "sector_result": sector_result,
    }

    proj = generate_projections(record, horizon_days=120, news_result=None, exclude_valuation=True)
    if proj.get("error"):
        meta["projection_error"] = proj["error"]
        return None, meta
    sig = proj.get("signal") or "NEUTRAL"
    meta["projection_signal"] = sig
    meta["composite_score"] = proj.get("composite_score")
    meta["confidence"] = proj.get("confidence")
    meta["ml_used"] = bool(proj.get("ml_used"))
    meta["p_up_5d"] = proj.get("p_up_5d")
    meta["p_up_20d"] = proj.get("p_up_20d")
    meta["p_up_60d"] = proj.get("p_up_60d")
    meta["p_up_120d"] = proj.get("p_up_120d")
    meta["ml_blend_weight"] = proj.get("ml_blend_weight_used")
    return projection_signal_to_classification(sig), meta


# ── run the full pipeline at a point in time ──────────────────────────────────

def classify_at(
    data: dict,
    raw: dict | None = None,
    cp_date: datetime | None = None,
    *,
    use_valuation: bool = True,
    signal_mode: str = "dcf",
    signal_meta: dict | None = None,
    spy_close_series: pd.Series | None = None,
) -> str | None:
    """Run the full engine pipeline on reconstructed data and return classification.

    ``signal_mode``:
      - ``dcf`` (default): ``classification_engine`` with optional DCF via ``use_valuation``.
      - ``tech_ai``: momentum + extended technicals + fundamentals (no DCF) +
        ``projection_engine.generate_projections(..., exclude_valuation=True)``;
        tier is derived from projection ``signal`` (BULLISH -> STRONG BUY, etc.).
        Requires ``raw`` (yfinance bundle) and ``cp_date`` (checkpoint).
    """
    signal_mode = normalize_signal_mode(signal_mode)
    if signal_mode == "tech_ai":
        if raw is None or cp_date is None:
            return None
        cl, meta = infer_classification_tech_ai(
            {**data}, raw, cp_date, spy_close_series=spy_close_series
        )
        if signal_meta is not None:
            signal_meta.update(meta)
        return cl

    try:
        sector_result = apply_sector_context(data)

        quality_result = analyze_quality(data)
        financial_result = analyze_financials(data)

        if use_valuation:
            valuation_result = analyze_valuation(
                {**data, "sector_result": sector_result},
                margin_of_safety=MARGIN_OF_SAFETY,
                wacc_adjustment=sector_result["wacc_adjustment"],
                terminal_growth_range=sector_result.get("terminal_growth_range"),
            )
            wacc = valuation_result.get("wacc_data", {}).get("wacc")
            valuation_metrics = valuation_result["valuation_metrics"]
            fair_value_weighted = valuation_result["fair_value_weighted"]
            buy_below_price = valuation_result["buy_below_price"]
            wacc_data = valuation_result["wacc_data"]
        else:
            wacc = None
            valuation_metrics = {}
            fair_value_weighted = None
            buy_below_price = None
            wacc_data = {}

        growth_result = analyze_growth(data)
        risk_result = analyze_risk(data)

        red_flag_result = analyze_red_flags(data, wacc=wacc)

        all_critical = (
            (financial_result.get("critical_flags") or []) +
            (risk_result.get("critical_flags") or [])
        )

        record = {
            **data,
            "valuation_metrics": valuation_metrics,
            "fair_value_weighted": fair_value_weighted,
            "buy_below_price": buy_below_price,
            "wacc_data": wacc_data,
            "red_flags": red_flag_result["red_flags"],
            "critical_flags": all_critical,
            "sector_result": sector_result,
        }

        clf_result = classify_stock(record)
        return clf_result["classification"]

    except Exception:
        logger.warning(
            "classify_at failed ticker=%s date=%s",
            data.get("ticker", "?"),
            cp_date,
            exc_info=True,
        )
        return None


# ── generate checkpoint dates ─────────────────────────────────────────────────

def _normalize_checkpoint_freq(freq: str | None) -> str:
    """Return ``Q`` (quarter-end) or ``M`` (month-end)."""
    if not freq:
        return "Q"
    x = str(freq).strip().lower()
    if x in ("m", "month", "monthly", "me", "mon"):
        return "M"
    return "Q"


def _generate_checkpoints(lookback_years: int = 2, freq: str | None = None) -> list[datetime]:
    """Checkpoint dates going back ``lookback_years`` from today.

    ``freq``: ``Q`` = quarter-ends (Mar/Jun/Sep/Dec), ``M`` = calendar month-ends.
    Last checkpoint is before ``today - max(FORWARD_PERIODS)`` so the longest forward return exists.
    """
    freq_u = _normalize_checkpoint_freq(freq)
    today = datetime.today()
    end = today - timedelta(days=int(max(FORWARD_PERIODS) * 30.44) + 15)
    start = today - timedelta(days=lookback_years * 365)

    if freq_u == "M":
        dr = pd.date_range(start=pd.Timestamp(start).normalize(), end=pd.Timestamp(end), freq="ME")
        checkpoints_m: list[datetime] = []
        for d in dr:
            dt = d.to_pydatetime()
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            if dt > start and dt <= end:
                checkpoints_m.append(dt)
        return sorted(set(checkpoints_m))

    checkpoints = []
    year = start.year
    quarter_ends = [
        datetime(year, 3, 31),
        datetime(year, 6, 30),
        datetime(year, 9, 30),
        datetime(year, 12, 31),
    ]

    current = start
    while current < end:
        for qe in quarter_ends:
            if qe > start and qe < end:
                checkpoints.append(qe)
        year += 1
        quarter_ends = [
            datetime(year, 3, 31),
            datetime(year, 6, 30),
            datetime(year, 9, 30),
            datetime(year, 12, 31),
        ]
        if quarter_ends[0] > today:
            break

    checkpoints = sorted(set(checkpoints))
    return checkpoints


# ── main backtest function ────────────────────────────────────────────────────

def run_backtest(
    tickers: list[str] | None = None,
    lookback_years: int = DEFAULT_LOOKBACK,
    *,
    benchmark: str | None = "SPY",
    yearly_top100: bool = False,
    universe_cache_dir: Path | None = None,
    universe_source: str = "pit",
    checkpoint_min_year: int | None = None,
    auto_build_missing_universe: bool = False,
    use_valuation: bool = True,
    signal_mode: str = "dcf",
    checkpoint_freq: str | None = None,
    hold_months: int = DEFAULT_HOLD_MONTHS,
) -> dict:
    signal_mode = normalize_signal_mode(signal_mode)
    """
    Run the full strategy backtest.

    When ``yearly_top100`` is True, ``tickers`` is ignored. For each checkpoint
    in calendar year C, only symbols listed for lag year C-1 are evaluated
    (see yearly_top100_universe).

    Returns dict with:
      - signals:    list of rows including optional ``universe_lag_year``
      - by_tier, summary, checkpoints, benchmark as before
      - yearly_top100, universe_by_lag_year (symbol counts), universe_cache_dir
      - use_valuation: whether DCF / valuation_engine was used (dcf mode only)
      - signal_mode: ``dcf`` | ``tech_ai`` (technicals + projection without DCF weight)
      - checkpoint_freq: ``Q`` (quarter-end) or ``M`` (month-end); more dates = more signals per ticker
    """
    cf = _normalize_checkpoint_freq(checkpoint_freq)
    uni_src: UniverseSource = (
        normalize_universe_source(universe_source) if yearly_top100 else "legacy"
    )

    if checkpoint_min_year is not None:
        y_now = datetime.today().year
        lookback_years = max(lookback_years, y_now - checkpoint_min_year + 2)

    checkpoints = _generate_checkpoints(lookback_years, cf)
    if checkpoint_min_year is not None:
        checkpoints = [c for c in checkpoints if c.year >= checkpoint_min_year]

    bench_hist = _load_benchmark_history(benchmark) if benchmark else None
    bench_label = (benchmark or "").upper() if bench_hist is not None else None

    udir = universe_cache_dir or default_universe_cache_dir(_ROOT, uni_src)
    universe_by_lag_year: dict[int, list[str]] | None = None
    if yearly_top100:
        if not checkpoints:
            print("  No checkpoints after filters — widen lookback or lower --checkpoint-min-year.")
            return {
                "signals": [],
                "by_tier": {},
                "summary": {"total_signals": 0, "verdict": "No checkpoints"},
                "checkpoints": [],
                "tickers": [],
                "benchmark": bench_label,
                "yearly_top100": True,
                "universe_source": uni_src,
                "universe_by_lag_year": {},
                "universe_map": None,
                "universe_cache_dir": str(udir),
                "use_valuation": use_valuation,
                "signal_mode": signal_mode,
                "checkpoint_freq": cf,
            }
        lag_years = sorted({c.year - 1 for c in checkpoints})
        universe_by_lag_year = load_universe_map_for_lag_years(
            lag_years,
            udir,
            auto_build_missing=auto_build_missing_universe,
            verbose=True,
            universe_source=uni_src,
        )
        all_tickers = sorted({t for y in lag_years for t in universe_by_lag_year.get(y, [])})
        if not all_tickers:
            raise ValueError(
                "Yearly top-100 mode produced an empty ticker union. "
                "Build universe files, e.g. "
                "python backtesting/build_yearly_top100_universe.py --from <y> --to <y>"
            )
    else:
        all_tickers = list(tickers if tickers is not None else DEFAULT_TICKERS)

    print(f"\n{'═' * 70}")
    print(f"  STRATEGY BACKTEST")
    print(f"{'═' * 70}")
    if yearly_top100:
        pool_label = "PIT S&P + dollar volume" if uni_src == "pit" else "current S&P 500 + dollar volume"
        print(f"  Mode:         Yearly top-100 (lag year = checkpoint year − 1)")
        print(f"  Universe:     {uni_src} ({pool_label})")
        print(f"  Universe dir: {udir}")
        print(f"  Union size:   {len(all_tickers)} distinct tickers across lag years")
        if checkpoints:
            from portfolio.universe_meta import pit_coverage

            cov = pit_coverage(
                checkpoints[0].date(),
                checkpoints[-1].date(),
            )
            if cov.get("pit_warning"):
                print(f"  PIT note:     {cov['pit_warning']}")
    else:
        print(f"  Tickers:      {len(all_tickers)} stocks")
    print(
        f"  Lookback:     {lookback_years} years ({len(checkpoints)} checkpoints, "
        f"{'month-end' if cf == 'M' else 'quarter-end'})"
    )
    if checkpoints:
        print(f"  Checkpoints:  {checkpoints[0].strftime('%Y-%m-%d')} -> {checkpoints[-1].strftime('%Y-%m-%d')}")
    if checkpoint_min_year is not None:
        print(f"  Min CP year:  {checkpoint_min_year}")
    print(f"  Forward:      {FORWARD_PERIODS} months (tier stats)")
    print(f"  Simulated hold: {hold_months} months (basket curve + dynamic backtest)")
    if bench_hist is not None:
        print(f"  Benchmark:    {bench_label} (excess = stock fwd − benchmark fwd)")
    else:
        print(f"  Benchmark:    (none)")
    signal_mode = normalize_signal_mode(signal_mode)
    if is_ml_strategy(signal_mode):
        print(f"  Valuation:    N/A (ML projection strategy, no DCF in composite)")
        print(f"  Signal mode:  {strategy_display_name(signal_mode)}")
    else:
        print(f"  Valuation:    {'DCF / valuation_engine ON' if use_valuation else 'OFF (quality / financials / risk / sector only)'}")
        print(f"  Signal mode:  {strategy_display_name(signal_mode)}")
    print(f"{'═' * 70}\n")

    signals = []
    spy_feat: pd.Series | None = None
    if bench_hist is not None and "Close" in bench_hist.columns:
        spy_feat = bench_hist["Close"].astype(float).copy()
        if getattr(spy_feat.index, "tz", None) is not None:
            spy_feat.index = spy_feat.index.tz_localize(None)

    for i, ticker in enumerate(all_tickers, 1):
        print(f"  [{i}/{len(all_tickers)}] {ticker} ... ", end="", flush=True)

        try:
            raw = collect_raw_yfinance(ticker)
        except Exception as e:
            print(f"FAILED (data error: {e})")
            continue

        if raw["income_statement"] is None:
            print("SKIPPED (no financials)")
            continue

        ticker_signals = 0
        for cp_date in checkpoints:
            if yearly_top100 and universe_by_lag_year is not None:
                ly = cp_date.year - 1
                allowed = set(universe_by_lag_year.get(ly, []))
                if ticker not in allowed:
                    continue

            data = reconstruct_data_at(raw, cp_date)
            if data is None:
                continue

            sig_meta: dict = {}
            classification = classify_at(
                data,
                raw,
                cp_date,
                use_valuation=use_valuation,
                signal_mode=signal_mode,
                signal_meta=sig_meta if is_ml_strategy(signal_mode) else None,
                spy_close_series=spy_feat,
            )
            if classification is None:
                continue

            # Measure forward returns
            price_at_cp = data["current_price"]
            fwd_returns = {}
            for months in FORWARD_PERIODS:
                fwd_price = _get_forward_price(raw["price_history"], cp_date, months)
                if fwd_price and price_at_cp and price_at_cp > 0:
                    fwd_returns[f"fwd_{months}m"] = (fwd_price - price_at_cp) / price_at_cp
                else:
                    fwd_returns[f"fwd_{months}m"] = None

            row: dict = {
                "ticker": ticker,
                "date": cp_date,
                "classification": classification,
                "price": price_at_cp,
                **fwd_returns,
            }
            if yearly_top100:
                row["universe_lag_year"] = cp_date.year - 1
            if is_ml_strategy(signal_mode):
                row["signal_mode"] = "ml"
                row["projection_signal"] = sig_meta.get("projection_signal")
                row["composite_score"] = sig_meta.get("composite_score")
                row["confidence"] = sig_meta.get("confidence")
                row["ml_used"] = sig_meta.get("ml_used")
                row["p_up_20d"] = sig_meta.get("p_up_20d")
                row["p_up_60d"] = sig_meta.get("p_up_60d")
                row["ml_score"] = ml_score_from_signal({**row, **sig_meta})

            if bench_hist is not None:
                spy_px = _get_price_at(bench_hist, cp_date)
                for months in FORWARD_PERIODS:
                    spy_fp = _get_forward_price(bench_hist, cp_date, months)
                    sk = f"spy_fwd_{months}m"
                    if spy_px and spy_fp and spy_px > 0:
                        row[sk] = (spy_fp - spy_px) / spy_px
                    else:
                        row[sk] = None
                    fk = f"fwd_{months}m"
                    ek = f"excess_fwd_{months}m"
                    fr, sr = row.get(fk), row.get(sk)
                    if fr is not None and sr is not None:
                        row[ek] = fr - sr
                    else:
                        row[ek] = None

            signals.append(row)
            ticker_signals += 1

        print(f"{ticker_signals} signals")

    # Aggregate results
    by_tier = _aggregate_by_tier(signals)
    summary = _compute_summary(signals, by_tier, hold_months=hold_months)

    uni_counts: dict[int, int] | None = None
    if universe_by_lag_year is not None:
        uni_counts = {y: len(v) for y, v in sorted(universe_by_lag_year.items())}

    ml_quintiles = None
    if is_ml_strategy(signal_mode):
        ml_quintiles = aggregate_quintile_forward_returns(signals, horizon_months=6)

    universe_meta: dict | None = None
    if yearly_top100 and checkpoints:
        from portfolio.universe_meta import universe_summary

        universe_meta = universe_summary(
            universe_source=uni_src,
            start=checkpoints[0].date(),
            end=checkpoints[-1].date(),
        )

    return {
        "signals": signals,
        "by_tier": by_tier,
        "summary": summary,
        "ml_quintiles": ml_quintiles,
        "checkpoints": checkpoints,
        "tickers": all_tickers,
        "benchmark": bench_label,
        "yearly_top100": yearly_top100,
        "universe_source": uni_src if yearly_top100 else None,
        "universe_meta": universe_meta,
        "universe_by_lag_year": uni_counts or {},
        "universe_map": universe_by_lag_year if yearly_top100 else None,
        "universe_cache_dir": str(udir),
        "use_valuation": use_valuation,
        "signal_mode": "ml" if is_ml_strategy(signal_mode) else MODE_DCF,
        "checkpoint_freq": cf,
        "hold_months": hold_months,
    }


# ── aggregation ───────────────────────────────────────────────────────────────

def _tier_stat_clip(r: float | None) -> float | None:
    """Winsorize one forward return for tier *average/median* stats (raw stays in Excel)."""
    if r is None:
        return None
    try:
        x = float(r)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    return max(-TIER_STATS_WINSOR, min(TIER_STATS_WINSOR, x))


def _aggregate_by_tier(signals: list[dict]) -> dict:
    """Compute average forward returns and hit rates by classification tier."""
    tiers = {}

    for tier in TIER_ORDER:
        tier_signals = [s for s in signals if s["classification"] == tier]
        if not tier_signals:
            continue

        tier_data = {"count": len(tier_signals), "signals": tier_signals}

        for period in FORWARD_PERIODS:
            key = f"fwd_{period}m"
            returns = [s[key] for s in tier_signals if s[key] is not None]
            returns_w = [_tier_stat_clip(s[key]) for s in tier_signals if s[key] is not None]
            if returns:
                n_raw = len(returns)
                n_eff = effective_n_for_signals(tier_signals, period, return_key=key)
                tier_data[f"avg_{period}m"] = sum(returns_w) / len(returns_w)
                tier_data[f"median_{period}m"] = sorted(returns_w)[len(returns_w) // 2]
                tier_data[f"hit_rate_{period}m"] = sum(1 for r in returns if r > 0) / n_raw
                tier_data[f"n_{period}m"] = n_raw
                tier_data[f"effective_n_{period}m"] = n_eff
                tier_data[f"overlap_factor_{period}m"] = overlap_inflation_factor(n_raw, n_eff)
                tier_data[f"best_{period}m"] = max(returns)
                tier_data[f"worst_{period}m"] = min(returns)
            else:
                tier_data[f"avg_{period}m"] = None
                tier_data[f"median_{period}m"] = None
                tier_data[f"hit_rate_{period}m"] = None
                tier_data[f"n_{period}m"] = 0
                tier_data[f"effective_n_{period}m"] = 0.0
                tier_data[f"overlap_factor_{period}m"] = None

        for period in FORWARD_PERIODS:
            ex_key = f"excess_fwd_{period}m"
            ex_returns = [s[ex_key] for s in tier_signals if s.get(ex_key) is not None]
            ex_w = [_tier_stat_clip(s.get(ex_key)) for s in tier_signals if s.get(ex_key) is not None]
            if ex_returns:
                tier_data[f"avg_excess_{period}m"] = sum(ex_w) / len(ex_w)
                tier_data[f"hit_rate_excess_{period}m"] = sum(1 for r in ex_returns if r > 0) / len(ex_returns)
                tier_data[f"n_excess_{period}m"] = len(ex_returns)
            else:
                tier_data[f"avg_excess_{period}m"] = None
                tier_data[f"hit_rate_excess_{period}m"] = None
                tier_data[f"n_excess_{period}m"] = 0

        tiers[tier] = tier_data

    return tiers


def _buy_avoid_spread(signals: list[dict], months: int) -> dict:
    """BUY vs AVOID average forward return spread for one horizon."""
    buy_tiers = ["STRONG BUY", "BUY"]
    avoid_tiers = ["AVOID", "STRONG AVOID"]
    fk = f"fwd_{months}m"
    exk = f"excess_fwd_{months}m"
    buy_r, avoid_r = [], []
    buy_ex, avoid_ex = [], []
    for s in signals:
        r = s.get(fk)
        if r is not None:
            if s["classification"] in buy_tiers:
                buy_r.append(r)
            elif s["classification"] in avoid_tiers:
                avoid_r.append(r)
        ex = s.get(exk)
        if ex is not None:
            if s["classification"] in buy_tiers:
                buy_ex.append(ex)
            elif s["classification"] in avoid_tiers:
                avoid_ex.append(ex)
    avg_buy = sum(_tier_stat_clip(r) for r in buy_r) / len(buy_r) if buy_r else None
    avg_avoid = sum(_tier_stat_clip(r) for r in avoid_r) / len(avoid_r) if avoid_r else None
    spread = (avg_buy - avg_avoid) if avg_buy is not None and avg_avoid is not None else None
    avg_buy_ex = sum(_tier_stat_clip(r) for r in buy_ex) / len(buy_ex) if buy_ex else None
    avg_avoid_ex = sum(_tier_stat_clip(r) for r in avoid_ex) / len(avoid_ex) if avoid_ex else None
    spread_ex = (
        (avg_buy_ex - avg_avoid_ex) if avg_buy_ex is not None and avg_avoid_ex is not None else None
    )
    return {
        f"avg_buy_{months}m": avg_buy,
        f"avg_avoid_{months}m": avg_avoid,
        f"buy_vs_avoid_spread_{months}m": spread,
        f"n_buy_signals_{months}m": len(buy_r),
        f"n_avoid_signals_{months}m": len(avoid_r),
        f"effective_n_buy_{months}m": effective_n_for_signals(
            [s for s in signals if s["classification"] in buy_tiers and s.get(fk) is not None],
            months,
        ),
        f"effective_n_avoid_{months}m": effective_n_for_signals(
            [s for s in signals if s["classification"] in avoid_tiers and s.get(fk) is not None],
            months,
        ),
        f"avg_buy_excess_{months}m": avg_buy_ex,
        f"avg_avoid_excess_{months}m": avg_avoid_ex,
        f"buy_vs_avoid_excess_spread_{months}m": spread_ex,
        f"n_buy_excess_signals_{months}m": len(buy_ex),
        f"n_avoid_excess_signals_{months}m": len(avoid_ex),
    }


def _compute_summary(signals: list[dict], by_tier: dict, *, hold_months: int = DEFAULT_HOLD_MONTHS) -> dict:
    """Compute overall backtest summary statistics."""
    total_signals = len(signals)
    if total_signals == 0:
        return {"total_signals": 0, "verdict": "No signals generated", "hold_months": hold_months}

    out: dict = {
        "total_signals": total_signals,
        "unique_tickers": len(set(s["ticker"] for s in signals)),
        "hold_months": hold_months,
        "distribution": {
            tier: sum(1 for s in signals if s["classification"] == tier)
            for tier in TIER_ORDER
            if sum(1 for s in signals if s["classification"] == tier) > 0
        },
    }
    for months in FORWARD_PERIODS:
        out.update(_buy_avoid_spread(signals, months))
        out[f"effective_n_{months}m_total"] = effective_n_for_signals(signals, months)

    # Legacy aliases for primary simulated hold (charts / verdict)
    hm = hold_months
    out["avg_buy_6m"] = out.get(f"avg_buy_{hm}m")
    out["avg_avoid_6m"] = out.get(f"avg_avoid_{hm}m")
    out["buy_vs_avoid_spread_6m"] = out.get(f"buy_vs_avoid_spread_{hm}m")
    out["n_buy_signals"] = out.get(f"n_buy_signals_{hm}m")
    out["n_avoid_signals"] = out.get(f"n_avoid_signals_{hm}m")
    out["effective_n_buy_6m"] = out.get(f"effective_n_buy_{hm}m")
    out["effective_n_avoid_6m"] = out.get(f"effective_n_avoid_{hm}m")
    out["avg_buy_excess_6m"] = out.get(f"avg_buy_excess_{hm}m")
    out["avg_avoid_excess_6m"] = out.get(f"avg_avoid_excess_{hm}m")
    out["buy_vs_avoid_excess_spread_6m"] = out.get(f"buy_vs_avoid_excess_spread_{hm}m")
    out["n_buy_excess_signals"] = out.get(f"n_buy_excess_signals_{hm}m")
    out["n_avoid_excess_signals"] = out.get(f"n_avoid_excess_signals_{hm}m")
    out["effective_n_6m_total"] = out.get(f"effective_n_{hm}m_total")
    return out


# ── display ───────────────────────────────────────────────────────────────────

def print_backtest_results(results: dict):
    """Print formatted backtest results to console."""
    by_tier = results["by_tier"]
    summary = results["summary"]

    print(f"\n{'═' * 70}")
    print(f"  BACKTEST RESULTS")
    print(f"{'═' * 70}")
    if is_ml_strategy(results.get("signal_mode")):
        print(f"  Signal mode:  {strategy_display_name(results.get('signal_mode'))}")
    if results.get("checkpoint_freq") == "M":
        print("  Checkpoints:  month-end (multiple evaluations per year per ticker)")
    if not results.get("use_valuation", True) and not is_ml_strategy(results.get("signal_mode")):
        print("  Valuation:    OFF (no DCF / valuation_engine in classifications)")
    print(f"  Total signals: {summary['total_signals']}  "
          f"({summary.get('unique_tickers', 0)} tickers × multiple checkpoints)")
    bench = results.get("benchmark")
    if bench:
        print(f"  Benchmark:     {bench} (excess = stock forward return − benchmark)")

    if summary["total_signals"] == 0:
        print(f"\n  No signals — nothing to aggregate.")
        print(f"{'═' * 70}")
        return

    # Distribution
    print(f"\n  Signal Distribution:")
    dist = summary.get("distribution", {})
    for tier in TIER_ORDER:
        if tier in dist:
            pct = dist[tier] / summary["total_signals"]
            bar = "█" * int(pct * 30)
            print(f"    {tier:<14} {dist[tier]:>4} signals ({pct:>5.1%})  {bar}")

    print(
        f"\n  Note: Avg/Median/Xs tier columns use winsorized returns (±{TIER_STATS_WINSOR:.0%}); "
        f"Best/Worst/Hit use raw. STRONG AVOID can look huge when *raw* outliers skew averages — "
        f"see Best 6M (e.g. one bad quote)."
    )
    hm = int(results.get("hold_months") or summary.get("hold_months") or DEFAULT_HOLD_MONTHS)
    eff_tot = summary.get(f"effective_n_{hm}m_total")
    if eff_tot is not None and summary["total_signals"] > 0:
        print(
            f"  Overlap:      raw signals={summary['total_signals']}, "
            f"effective_n({hm}M)≈{eff_tot:.0f} "
            f"(non-overlapping forward windows per ticker; use for inference, not raw N)"
        )
    print(f"  Simulated hold: {hm} months (basket curve / dynamic backtest; tier table still shows 3/6/12M)")

    # Returns by tier
    if bench:
        print(f"\n  {'─' * 78}")
        print(f"  {'TIER':<14} {'Count':>6} │ {'Avg 3M':>8} {'Avg 6M':>8} {'Avg 12M':>8} │ {'Hit 6M':>7} │ {'Xs 6M':>8}")
        print(f"  {'─' * 78}")
    else:
        print(f"\n  {'─' * 66}")
        print(f"  {'TIER':<14} {'Count':>6} │ {'Avg 3M':>8} {'Avg 6M':>8} {'Avg 12M':>8} │ {'Hit 6M':>7}")
        print(f"  {'─' * 66}")

    for tier in TIER_ORDER:
        if tier not in by_tier:
            continue
        t = by_tier[tier]
        avg_3 = f"{t['avg_3m']:+.1%}" if t.get("avg_3m") is not None else "  N/A "
        avg_6 = f"{t['avg_6m']:+.1%}" if t.get("avg_6m") is not None else "  N/A "
        avg_12 = f"{t['avg_12m']:+.1%}" if t.get("avg_12m") is not None else "  N/A "
        hit_6 = f"{t['hit_rate_6m']:.0%}" if t.get("hit_rate_6m") is not None else " N/A "
        n_eff = t.get("effective_n_6m")
        eff_s = f" (eff≈{n_eff:.0f})" if n_eff is not None and t.get("n_6m") else ""
        if bench:
            xs = t.get("avg_excess_6m")
            xs_s = f"{xs:+.1%}" if xs is not None else "  N/A "
            print(
                f"  {tier:<14} {t['count']:>6}{eff_s} │ {avg_3:>8} {avg_6:>8} {avg_12:>8} │ {hit_6:>7} │ {xs_s:>8}"
            )
        else:
            print(f"  {tier:<14} {t['count']:>6}{eff_s} │ {avg_3:>8} {avg_6:>8} {avg_12:>8} │ {hit_6:>7}")

    if bench:
        print(f"  {'─' * 78}")
    else:
        print(f"  {'─' * 66}")

    # Key verdict — all forward horizons; primary simulated hold = hm
    print(f"\n  KEY FINDINGS (BUY vs AVOID spread by forward horizon):")
    for months in FORWARD_PERIODS:
        sp = summary.get(f"buy_vs_avoid_spread_{months}m")
        if sp is not None:
            print(f"    • {months}M spread: {sp:+.1%}  "
                  f"(n_buy={summary.get(f'n_buy_signals_{months}m', 0)})")

    spread = summary.get(f"buy_vs_avoid_spread_{hm}m")
    if spread is not None:
        direction = "outperformed" if spread > 0 else "underperformed"
        if spread > 0.05:
            print(f"\n  VERDICT ({hm}M simulated hold): BUY {direction} AVOID by {abs(spread):.1%}.")
        elif spread > 0:
            print(f"\n  VERDICT ({hm}M hold): weak positive — BUY marginally beat AVOID.")
        else:
            print(f"\n  VERDICT ({hm}M hold): signal inversion — AVOID beat BUY.")
    else:
        print(f"\n  VERDICT: Insufficient BUY/AVOID signals to compute spread.")

    if is_ml_strategy(results.get("signal_mode")) and results.get("ml_quintiles"):
        print_quintile_table(results["ml_quintiles"], horizon_months=6)
        print(
            "  Tip: if Q5−Q1 is positive, use market_neutral_backtest.py (long top / short bottom by score)."
        )

    if results.get("benchmark") and summary.get("avg_buy_excess_6m") is not None:
        b = results["benchmark"]
        print(f"\n  VS {b} (6M excess on overlapping signals):")
        print(f"    • BUY-tier avg excess:      {summary['avg_buy_excess_6m']:+.1%}  "
              f"({summary.get('n_buy_excess_signals', 0)} signals)")
        if summary.get("avg_avoid_excess_6m") is not None:
            print(f"    • AVOID-tier avg excess:    {summary['avg_avoid_excess_6m']:+.1%}  "
                  f"({summary.get('n_avoid_excess_signals', 0)} signals)")
        sx = summary.get("buy_vs_avoid_excess_spread_6m")
        if sx is not None:
            print(f"    • BUY vs AVOID excess spread: {sx:+.1%}")

    print(f"{'═' * 70}")

    # Per-ticker breakdown
    _print_ticker_summary(results["signals"])


def _print_ticker_summary(signals: list[dict]):
    """Print per-ticker signal summary."""
    tickers = sorted(set(s["ticker"] for s in signals))
    if not tickers:
        return

    print(f"\n  PER-TICKER BREAKDOWN:")
    print(f"  {'Ticker':<10} {'Signals':>8} {'Most Common':>14} {'Avg 6M Fwd':>12} {'Best':>8} {'Worst':>8}")
    print(f"  {'─' * 62}")

    for ticker in tickers:
        t_signals = [s for s in signals if s["ticker"] == ticker]
        classifications = [s["classification"] for s in t_signals]
        most_common = max(set(classifications), key=classifications.count)

        returns_6m = [s["fwd_6m"] for s in t_signals if s["fwd_6m"] is not None]
        avg_6m = sum(returns_6m) / len(returns_6m) if returns_6m else None
        best = max(returns_6m) if returns_6m else None
        worst = min(returns_6m) if returns_6m else None

        avg_str = f"{avg_6m:+.1%}" if avg_6m is not None else "N/A"
        best_str = f"{best:+.1%}" if best is not None else "N/A"
        worst_str = f"{worst:+.1%}" if worst is not None else "N/A"

        print(f"  {ticker:<10} {len(t_signals):>8} {most_common:>14} {avg_str:>12} {best_str:>8} {worst_str:>8}")


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(results: dict, filepath: str | None = None) -> str:
    """Export backtest results to an Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl not installed — skipping Excel export")
        return ""

    if filepath is None:
        filepath = str(_ROOT / f"backtest_results_{datetime.today().strftime('%Y%m%d')}.xlsx")

    wb = Workbook()

    # Sheet 1: Summary by tier
    ws = wb.active
    ws.title = "By Classification"
    bench = results.get("benchmark")
    headers = [
        "Tier", "Signals", "Eff N 6M", "Avg 3M", "Avg 6M", "Avg 12M",
        "Hit Rate 6M", "Best 6M", "Worst 6M",
    ]
    if bench:
        headers += ["Avg Excess 6M", "Hit Excess 6M"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    row = 2
    for tier in TIER_ORDER:
        if tier not in results["by_tier"]:
            continue
        t = results["by_tier"][tier]
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=t["count"])
        ws.cell(row=row, column=3, value=t.get("effective_n_6m"))
        ws.cell(row=row, column=4, value=t.get("avg_3m"))
        ws.cell(row=row, column=5, value=t.get("avg_6m"))
        ws.cell(row=row, column=6, value=t.get("avg_12m"))
        ws.cell(row=row, column=7, value=t.get("hit_rate_6m"))
        ws.cell(row=row, column=8, value=t.get("best_6m"))
        ws.cell(row=row, column=9, value=t.get("worst_6m"))
        last_pct_col = 9
        if bench:
            ws.cell(row=row, column=10, value=t.get("avg_excess_6m"))
            ws.cell(row=row, column=11, value=t.get("hit_rate_excess_6m"))
            last_pct_col = 11
        for col in range(4, last_pct_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = "0.0%"
        row += 1

    # Sheet 2: All signals
    ws2 = wb.create_sheet("All Signals")
    has_univ = any("universe_lag_year" in s for s in results["signals"])
    headers2 = ["Ticker", "Date", "Classification"]
    if has_univ:
        headers2.append("Univ lag Yr")
    has_ml = any(s.get("p_up_20d") is not None for s in results["signals"])
    headers2 += ["Price", "Fwd 3M", "Fwd 6M", "Fwd 12M"]
    if has_ml:
        headers2 += ["Proj signal", "P(up) 20d", "P(up) 60d", "ML used"]
    if bench:
        headers2 += ["Spy Fwd 3M", "Spy Fwd 6M", "Spy Fwd 12M", "Excess 3M", "Excess 6M", "Excess 12M"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h).font = Font(bold=True)

    for row_idx, s in enumerate(results["signals"], 2):
        col = 1
        ws2.cell(row=row_idx, column=col, value=s["ticker"])
        col += 1
        ws2.cell(row=row_idx, column=col, value=s["date"].strftime("%Y-%m-%d"))
        col += 1
        ws2.cell(row=row_idx, column=col, value=s["classification"])
        col += 1
        if has_univ:
            ws2.cell(row=row_idx, column=col, value=s.get("universe_lag_year"))
            col += 1
        ws2.cell(row=row_idx, column=col, value=s["price"])
        col += 1
        c_fwd3 = col
        ws2.cell(row=row_idx, column=col, value=s.get("fwd_3m"))
        col += 1
        c_fwd6 = col
        ws2.cell(row=row_idx, column=col, value=s.get("fwd_6m"))
        col += 1
        c_fwd12 = col
        ws2.cell(row=row_idx, column=col, value=s.get("fwd_12m"))
        col += 1
        pct_cols = [c_fwd3, c_fwd6, c_fwd12]
        if has_ml:
            ws2.cell(row=row_idx, column=col, value=s.get("projection_signal"))
            col += 1
            c_p20 = col
            ws2.cell(row=row_idx, column=col, value=s.get("p_up_20d"))
            col += 1
            c_p60 = col
            ws2.cell(row=row_idx, column=col, value=s.get("p_up_60d"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("ml_used"))
            col += 1
            pct_cols.extend([c_p20, c_p60])
        if bench:
            c0 = col
            ws2.cell(row=row_idx, column=col, value=s.get("spy_fwd_3m"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("spy_fwd_6m"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("spy_fwd_12m"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("excess_fwd_3m"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("excess_fwd_6m"))
            col += 1
            ws2.cell(row=row_idx, column=col, value=s.get("excess_fwd_12m"))
            col += 1
            pct_cols.extend(range(c0, c0 + 6))
        for pc in pct_cols:
            cell = ws2.cell(row=row_idx, column=pc)
            if cell.value is not None:
                cell.number_format = "0.0%"

    wb.save(filepath)
    return filepath


def export_backtest_vs_benchmark_html(
    results: dict,
    filepath: str | None = None,
    *,
    portfolio_weight_mode: str = "tier",
) -> str:
    """
    Write a standalone Plotly HTML chart: per-signal excess vs benchmark, quarterly averages,
    and a sequential **weighted BUY basket** vs SPY (non-overlapping holds of ``hold_months``).
    """
    hm = int(results.get("hold_months") or DEFAULT_HOLD_MONTHS)
    exk = f"excess_fwd_{hm}m"
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print("  plotly not installed — skipping HTML chart export")
        return ""

    bench = results.get("benchmark")
    if not bench:
        return ""

    signals = [s for s in results["signals"] if s.get(exk) is not None]
    if not signals:
        print(f"  No {hm}M excess vs benchmark — skipping HTML chart (need SPY data at each signal)")
        return ""

    if filepath is None:
        filepath = str(_ROOT / f"backtest_vs_{bench.lower()}_{datetime.today().strftime('%Y%m%d')}.html")

    tier_colors = {
        "STRONG BUY": "#2ecc71",
        "BUY": "#58d68d",
        "WATCHLIST": "#f4d03f",
        "HOLD": "#95a5a6",
        "AVOID": "#e74c3c",
        "STRONG AVOID": "#c0392b",
    }

    fig = make_subplots(
        rows=3,
        cols=1,
        shared_xaxes=False,
        vertical_spacing=0.10,
        subplot_titles=(
            f"{hm}-month forward excess vs {bench} (each point = one ticker at one checkpoint)",
            f"Average {hm}M excess vs {bench} by signal quarter",
            f"Cumulative $1 — sequential BUY basket vs {bench} ({hm}M non-overlapping holds; {portfolio_weight_mode} weights)",
        ),
        row_heights=[0.36, 0.28, 0.36],
    )

    for tier in TIER_ORDER:
        sub = [s for s in signals if s.get("classification") == tier]
        if not sub:
            continue
        fig.add_trace(
            go.Scatter(
                x=[s["date"] for s in sub],
                y=[s[exk] for s in sub],
                mode="markers",
                name=tier,
                marker=dict(color=tier_colors.get(tier, "#3498db"), size=7, opacity=0.75),
                text=[f"{s['ticker']} @ {s['date'].strftime('%Y-%m-%d')}" for s in sub],
                hovertemplate=f"%{{text}}<br>Excess {hm}M: %{{y:.2%}}<extra></extra>",
            ),
            row=1,
            col=1,
        )

    fig.add_hline(y=0, line_dash="dash", line_color="white", opacity=0.5, row=1, col=1)

    df = pd.DataFrame(
        {"date": [s["date"] for s in signals], "excess": [s[exk] for s in signals]}
    )
    df["date"] = pd.to_datetime(df["date"])
    df["q_end"] = df["date"].dt.to_period("Q").dt.to_timestamp(how="end")
    qmean = df.groupby("q_end", as_index=False)["excess"].mean()

    fig.add_trace(
        go.Bar(
            x=qmean["q_end"],
            y=qmean["excess"],
            name="Quarter avg excess",
            marker_color="#5dade2",
            hovertemplate=f"Quarter ending %{{x|%Y-%m-%d}}<br>Mean excess {hm}M: %{{y:.2%}}<extra></extra>",
        ),
        row=2,
        col=1,
    )
    fig.add_hline(y=0, line_dash="dash", line_color="white", opacity=0.5, row=2, col=1)

    curve = sequential_weighted_equity_curve(
        results["signals"],
        weight_mode=portfolio_weight_mode if portfolio_weight_mode in ("tier", "equal") else "tier",
        hold_months=hm,
    )
    if not curve.empty:
        fig.add_trace(
            go.Scatter(
                x=curve["exit_date"],
                y=curve["equity_stock"],
                mode="lines+markers",
                name="BUY basket (strategy)",
                line=dict(color="#2ecc71", width=2),
                marker=dict(size=8),
                hovertemplate="Exit %{x|%Y-%m-%d}<br>Equity: %{y:.3f}<extra></extra>",
            ),
            row=3,
            col=1,
        )
        fig.add_trace(
            go.Scatter(
                x=curve["exit_date"],
                y=curve["equity_spy"],
                mode="lines+markers",
                name=f"Same basket × {bench}",
                line=dict(color="#aeb6bf", width=2, dash="dot"),
                marker=dict(size=8),
                hovertemplate="Exit %{x|%Y-%m-%d}<br>Equity: %{y:.3f}<extra></extra>",
            ),
            row=3,
            col=1,
        )
        fig.add_hline(y=1.0, line_dash="dash", line_color="white", opacity=0.4, row=3, col=1)

    fig.update_layout(
        template="plotly_dark",
        height=1020,
        title=dict(
            text=f"Backtest vs {bench} (excess = stock 6M forward − {bench} 6M forward)",
            x=0.02,
            xanchor="left",
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(t=100, b=60),
    )
    fig.update_yaxes(tickformat=".0%", row=1, col=1)
    fig.update_yaxes(tickformat=".0%", row=2, col=1)
    fig.update_yaxes(tickformat=".3f", row=3, col=1)
    fig.update_xaxes(title_text="Checkpoint date", row=1, col=1)
    fig.update_xaxes(title_text="Quarter (end)", row=2, col=1)
    fig.update_xaxes(title_text="Hold exit date (end of each 6M window)", row=3, col=1)

    fig.write_html(filepath, include_plotlyjs="cdn", config={"displayModeBar": True})
    return filepath


# ── CLI entry point ───────────────────────────────────────────────────────────

DEFAULT_TICKERS = [
    "MAIN", "BTI", "SHEL", "O", "MNG.L", "AGNC", "BMO", "BNS",
    "BMW.DE", "CNQ", "MBG.DE", "PFE", "RIO", "STAG", "TD", "UPS",
]


def main():
    tickers_from_file: list[str] | None = None
    extra_positional: list[str] = []
    lookback = DEFAULT_LOOKBACK
    benchmark = "SPY"
    yearly_top100 = False
    explicit_yearly_top100 = False
    classic_tickers = False
    checkpoint_min_year: int | None = None
    universe_dir: Path | None = None
    universe_source = "pit"
    auto_build_universe = False
    portfolio_weight = "tier"
    legacy_html = False
    use_valuation = True
    signal_mode = "dcf"
    checkpoint_freq: str | None = None
    hold_months = DEFAULT_HOLD_MONTHS
    run_market_neutral_flag = False

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--lookback" and i + 1 < len(args):
            lookback = int(args[i + 1])
            i += 2
        elif a == "--benchmark" and i + 1 < len(args):
            benchmark = args[i + 1]
            i += 2
        elif a == "--tickers-file" and i + 1 < len(args):
            tickers_from_file = _load_tickers_from_file(args[i + 1])
            i += 2
        elif a == "--yearly-top100":
            yearly_top100 = True
            explicit_yearly_top100 = True
            i += 1
        elif a == "--classic-tickers":
            classic_tickers = True
            yearly_top100 = False
            i += 1
        elif a == "--checkpoint-min-year" and i + 1 < len(args):
            checkpoint_min_year = int(args[i + 1])
            i += 2
        elif a == "--universe-dir" and i + 1 < len(args):
            universe_dir = Path(args[i + 1])
            i += 2
        elif a == "--auto-build-universe":
            auto_build_universe = True
            i += 1
        elif a == "--universe-source" and i + 1 < len(args):
            universe_source = args[i + 1]
            i += 2
        elif a == "--portfolio-weight" and i + 1 < len(args):
            portfolio_weight = args[i + 1].strip().lower()
            i += 2
        elif a == "--legacy-html":
            legacy_html = True
            i += 1
        elif a == "--no-valuation":
            use_valuation = False
            i += 1
        elif a == "--signal-tech-ai":
            signal_mode = "ml"
            i += 1
        elif a == "--strategy" and i + 1 < len(args):
            signal_mode = args[i + 1]
            i += 2
        elif a == "--market-neutral":
            run_market_neutral_flag = True
            i += 1
        elif a == "--checkpoint-freq" and i + 1 < len(args):
            checkpoint_freq = args[i + 1]
            i += 2
        elif a == "--hold-months" and i + 1 < len(args):
            hold_months = int(args[i + 1])
            i += 2
        elif a == "--help":
            print(__doc__)
            return
        elif a.startswith("--"):
            print(f"Unknown option: {a}\n{__doc__}")
            return
        else:
            extra_positional.append(a)
            i += 1

    if (
        not classic_tickers
        and not explicit_yearly_top100
        and tickers_from_file is None
        and not extra_positional
    ):
        pit_dir = default_universe_cache_dir(_ROOT, "pit")
        leg_dir = default_universe_cache_dir(_ROOT, "legacy")
        if universe_dir is not None:
            u_check = universe_dir
        elif pit_dir.is_dir() and any(pit_dir.glob("*.txt")):
            u_check = pit_dir
            universe_source = "pit"
        elif leg_dir.is_dir() and any(leg_dir.glob("*.txt")):
            u_check = leg_dir
            universe_source = "legacy"
        else:
            u_check = pit_dir
        if u_check.is_dir() and any(u_check.glob("*.txt")):
            yearly_top100 = True
        else:
            print(
                "\n  No yearly top-100 universe cache found (expected .txt per year under "
                f"{u_check}).\n"
                "  Falling back to the small classic ticker demo list.\n"
                "  To use top-100: run  python backtesting/build_yearly_top100_universe.py "
                "--for-checkpoints-from-year 2023 --universe-source pit\n"
            )
            classic_tickers = True

    if yearly_top100 and (tickers_from_file is not None or extra_positional):
        print(
            "  Note: --yearly-top100 ignores --tickers-file and extra ticker arguments "
            "(universe comes from cached yearly lists).\n"
        )

    if tickers_from_file is not None and not yearly_top100:
        merged = list(
            dict.fromkeys(
                [t.strip().upper() for t in tickers_from_file]
                + [t.strip().upper() for t in extra_positional]
            )
        )
        tickers = merged
    elif extra_positional and not yearly_top100:
        tickers = [t.strip().upper() for t in extra_positional]
    else:
        tickers = None if yearly_top100 else DEFAULT_TICKERS

    bm_arg: str | None = benchmark
    if benchmark and benchmark.lower() in ("none", "off", "false", "-"):
        bm_arg = None

    if portfolio_weight not in ("tier", "equal"):
        print(f"  Unknown --portfolio-weight {portfolio_weight!r}, using 'tier'.")
        portfolio_weight = "tier"

    if yearly_top100 and checkpoint_min_year is None:
        checkpoint_min_year = DEFAULT_TOP100_CHECKPOINT_MIN_YEAR

    signal_mode = normalize_signal_mode(signal_mode)

    print(f"\nStrategy Backtest — {datetime.today().strftime('%Y-%m-%d')}")

    try:
        results = run_backtest(
            tickers,
            lookback_years=lookback,
            benchmark=bm_arg,
            yearly_top100=yearly_top100,
            universe_cache_dir=universe_dir,
            universe_source=universe_source,
            checkpoint_min_year=checkpoint_min_year,
            auto_build_missing_universe=auto_build_universe,
            use_valuation=use_valuation,
            signal_mode=signal_mode,
            checkpoint_freq=checkpoint_freq,
            hold_months=hold_months,
        )
    except FileNotFoundError as e:
        print(f"\n  {e}")
        print(
            "\n  Yearly top-100 mode needs universe cache files. Build them, for example:\n"
            "    python backtesting/build_yearly_top100_universe.py "
            "--for-checkpoints-from-year 2023 --universe-source pit\n"
            "  Or run with an explicit ticker list / --classic-tickers.\n"
        )
        return
    print_backtest_results(results)

    # Excel export
    try:
        path = export_to_excel(results)
        if path:
            print(f"\n  Excel report saved -> {path}")
    except Exception as e:
        print(f"\n  Excel export failed: {e}")

    if bm_arg and results.get("tickers"):
        try:
            from backtesting.dynamic_portfolio_backtest import run_dynamic

            dyn_suffix = "ml" if is_ml_strategy(results.get("signal_mode")) else "dcf"
            dyn_path = _ROOT / f"strategy_dynamic_{dyn_suffix}_vs_spy_{datetime.today().strftime('%Y%m%d')}.html"
            run_dynamic(
                lookback_years=lookback,
                checkpoint_min_year=checkpoint_min_year,
                universe_dir=universe_dir,
                auto_build_universe=auto_build_universe,
                breakout_days=20,
                stop_loss=0.05,
                take_profit=0.25,
                max_hold_days=hold_days(hold_months),
                universe_source=universe_source,
                position_frac=0.10,
                max_positions=10,
                max_tickers=None,
                out_html=dyn_path,
                tickers=list(results["tickers"]),
                universe_map=results.get("universe_map"),
                use_valuation=use_valuation,
                signal_mode=results.get("signal_mode", "dcf"),
                checkpoint_freq=results.get("checkpoint_freq", "Q"),
                min_p_up_20d=0.52 if is_ml_strategy(results.get("signal_mode")) else None,
                regime_filter=is_ml_strategy(results.get("signal_mode")),
                entry_mode="rank" if is_ml_strategy(results.get("signal_mode")) else "tier",
            )
            print(f"\n  Dynamic strategy vs SPY buy-and-hold (animated monthly) -> {dyn_path}")
            print("     Use the slider / Play control to step through months.")
        except Exception as e:
            print(f"\n  Dynamic HTML export failed: {e}")

    if run_market_neutral_flag and is_ml_strategy(results.get("signal_mode")):
        try:
            from backtesting.market_neutral_backtest import run_market_neutral_from_results

            mn = run_market_neutral_from_results(results)
            print(f"  Market-neutral L/S chart -> {mn['html']}")
        except Exception as e:
            print(f"\n  Market-neutral backtest failed: {e}")

    if bm_arg and legacy_html:
        try:
            hpath = export_backtest_vs_benchmark_html(results, portfolio_weight_mode=portfolio_weight)
            if hpath:
                print(f"\n  Legacy diagnostic chart -> {hpath}")
        except Exception as e:
            print(f"\n  Legacy HTML export failed: {e}")


if __name__ == "__main__":
    main()
