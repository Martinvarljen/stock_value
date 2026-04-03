from __future__ import annotations

"""
excel_output.py  —  Export analysis results to a formatted Excel workbook

write_excel(results, filepath) → saves workbook to filepath

Sheets:
  1. Summary      — one row per ticker, all key metrics at a glance
  2. Valuation    — DCF details, fair value, buy-below, multiples
  3. Quality      — margins, ROIC, growth, dilution
  4. Financials   — leverage, coverage, FCF strength
  5. Risk         — beta, volatility, critical flags
  6. Red Flags    — all detected patterns with severity

Requires: openpyxl  (pip install openpyxl)
"""

from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from utils import _pct, _num, _x, _bn
from peer_engine import analyze_peers


# ── colour palette ─────────────────────────────────────────────────────────────
_C = {
    "header_dark":  "1F3864",   # dark navy
    "header_mid":   "2E5D9E",   # mid blue
    "header_light": "4472C4",   # light blue
    "strong_buy":   "1E7B34",   # dark green
    "buy":          "70AD47",   # green
    "watchlist":    "92D050",   # light green
    "hold":         "FFD966",   # amber
    "avoid":        "FF7043",   # orange-red
    "strong_avoid": "C00000",   # dark red
    "high_rf":      "FFCCCC",   # pale red
    "medium_rf":    "FFF2CC",   # pale amber
    "low_rf":       "EAF1FB",   # pale blue
    "row_alt":      "F2F7FF",   # very light blue for alternating rows
    "white":        "FFFFFF",
    "black":        "000000",
    "border":       "B8CCE4",
}

_CLASSIFICATION_COLORS = {
    "STRONG BUY":   _C["strong_buy"],
    "BUY":          _C["buy"],
    "WATCHLIST":    _C["watchlist"],
    "HOLD":         _C["hold"],
    "AVOID":        _C["avoid"],
    "STRONG AVOID": _C["strong_avoid"],
}

_SEVERITY_COLORS = {
    "HIGH":   _C["high_rf"],
    "MEDIUM": _C["medium_rf"],
    "LOW":    _C["low_rf"],
}


# ── style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    # Explicit font name so output matches the spec better.
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _border() -> Border:
    thin = Side(style="thin", color=_C["border"])
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(cell, level: int = 1):
    """Level 1 = dark navy, 2 = mid blue, 3 = light blue."""
    colors = {1: _C["header_dark"], 2: _C["header_mid"], 3: _C["header_light"]}
    cell.fill    = _fill(colors.get(level, _C["header_light"]))
    cell.font    = _font(bold=True, color=_C["white"], size=10)
    cell.alignment = _center()
    cell.border  = _border()


def _style_data(cell, alt_row=False, bold=False, align="left"):
    cell.fill      = _fill(_C["row_alt"] if alt_row else _C["white"])
    cell.font      = _font(bold=bold, size=10)
    cell.alignment = _left() if align == "left" else _center()
    cell.border    = _border()


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row: int, values: list, alt=False, bold=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        _style_data(cell, alt_row=alt, bold=bold,
                    align="center" if c > 1 else "left")


# ── sheet builders ─────────────────────────────────────────────────────────────

def _build_summary(wb: "Workbook", results: list[dict]):
    ws = wb.create_sheet("Summary", 0)
    ws.freeze_panes = "C3"

    title_cell = ws.cell(row=1, column=1,
                         value=f"Stock Analyzer — {datetime.today().strftime('%Y-%m-%d')}")
    title_cell.font      = _font(bold=True, size=14, color=_C["header_dark"])
    title_cell.alignment = _left()
    ws.row_dimensions[1].height = 22

    headers = [
        "Ticker", "Company", "Sector", "Currency",
        "Price", "EUR Price", "Fair Value (DCF)", "Buy-Below",
        "Upside %", "Classification",
        "ROIC", "WACC", "ROIC/WACC Spread",
        "Rev CAGR 5Y", "Op Margin",
        "Net Debt/EBITDA", "Interest Coverage",
        "Red Flags (H/M/L)", "Data Quality",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        _style_header(cell, level=1)
    ws.row_dimensions[2].height = 30

    for i, r in enumerate(results):
        row    = i + 3
        alt    = (i % 2 == 1)
        price  = r.get("current_price")
        fv     = r.get("fair_value_weighted")
        bb     = r.get("buy_below_price")
        upside = ((fv - price) / price) if (fv and price and price > 0) else None
        clf    = (r.get("classification_result") or {}).get("classification", "N/A")
        roic   = r.get("roic")
        wacc   = (r.get("wacc_data") or {}).get("wacc")
        spread = (roic - wacc) if (roic is not None and wacc is not None) else None
        rf     = r.get("red_flags") or []
        h_rf   = sum(1 for f in rf if f["severity"] == "HIGH")
        m_rf   = sum(1 for f in rf if f["severity"] == "MEDIUM")
        l_rf   = sum(1 for f in rf if f["severity"] == "LOW")

        values = [
            r.get("ticker", ""),
            (r.get("company_name") or "")[:40],
            (r.get("sector") or "")[:30],
            r.get("currency", ""),
            round(price, 2) if price else None,
            round(r.get("price_eur"), 2) if r.get("price_eur") else None,
            round(fv, 2) if fv else None,
            round(bb, 2) if bb else None,
            f"{upside:+.1%}" if upside is not None else "N/A",
            clf,
            f"{roic:.1%}" if roic is not None else "N/A",
            f"{wacc:.1%}" if wacc is not None else "N/A",
            f"{spread:+.1%}" if spread is not None else "N/A",
            f"{r.get('revenue_cagr_5y'):.1%}" if r.get("revenue_cagr_5y") is not None else "N/A",
            f"{r.get('operating_margin'):.1%}" if r.get("operating_margin") is not None else "N/A",
            f"{r.get('net_debt_ebitda'):.1f}x" if r.get("net_debt_ebitda") is not None else "N/A",
            f"{r.get('interest_coverage'):.1f}x" if r.get("interest_coverage") is not None else "N/A",
            f"{h_rf}H / {m_rf}M / {l_rf}L",
            f"{r.get('data_quality_score', 0)}/100",
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            _style_data(cell, alt_row=alt, align="center" if c > 1 else "left")

        # Colour the classification cell
        clf_cell = ws.cell(row=row, column=10)
        clf_color = _CLASSIFICATION_COLORS.get(clf)
        if clf_color:
            clf_cell.fill = _fill(clf_color)
            clf_cell.font = _font(bold=True,
                                  color=_C["white"] if clf in ("STRONG BUY", "STRONG AVOID", "AVOID", "BUY") else _C["black"])

        # Colour the upside cell
        upside_cell = ws.cell(row=row, column=9)
        if upside is not None:
            if upside > 0.15:
                upside_cell.fill = _fill(_C["buy"])
            elif upside < -0.15:
                upside_cell.fill = _fill(_C["avoid"])

        ws.row_dimensions[row].height = 18

    # Column widths
    widths = [10, 32, 22, 9, 9, 9, 14, 10, 9, 14,
              8, 8, 14, 12, 10, 15, 17, 16, 12]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)


def _build_detail_sheet(wb: "Workbook", r: dict, sheet_name: str):
    """One sheet per ticker with all metric blocks laid out in sections."""
    ws = wb.create_sheet(sheet_name[:31])

    # Title
    ticker  = r.get("ticker", "")
    company = r.get("company_name") or ticker
    price   = r.get("current_price")
    curr    = r.get("currency", "")
    clf     = (r.get("classification_result") or {}).get("classification", "N/A")

    price_label = f"{curr} {price:.2f}" if price is not None else "Price N/A"
    title = ws.cell(row=1, column=1,
                    value=f"{company} ({ticker})  |  {price_label}  |  {clf}")
    title.font      = _font(bold=True, size=13, color=_C["header_dark"])
    title.alignment = _left()
    ws.row_dimensions[1].height = 22
    ws.merge_cells("A1:F1")

    _set_col_width(ws, 1, 36)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 55)
    _set_col_width(ws, 5, 28)
    _set_col_width(ws, 6, 40)

    current_row = [3]   # mutable so helper can advance it

    def _section(title_text: str):
        r_idx = current_row[0]
        for c in range(1, 7):
            cell = ws.cell(row=r_idx, column=c)
            _style_header(cell, level=2)
        ws.cell(row=r_idx, column=1, value=title_text)
        ws.cell(row=r_idx, column=1).font = _font(bold=True, color=_C["white"], size=11)
        ws.merge_cells(f"A{r_idx}:F{r_idx}")
        ws.row_dimensions[r_idx].height = 20
        current_row[0] += 1

        # Column sub-headers
        sub_headers = ["Metric", "Value", "Formatted", "Assessment", "Benchmark", "Detail"]
        for c, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=current_row[0], column=c, value=h)
            _style_header(cell, level=3)
        ws.row_dimensions[current_row[0]].height = 16
        current_row[0] += 1

    def _metrics_block(metrics_dict: dict):
        if not metrics_dict:
            return
        for idx, (key, m) in enumerate(metrics_dict.items()):
            if not isinstance(m, dict):
                continue
            alt  = (idx % 2 == 1)
            r_idx = current_row[0]
            vals = [
                m.get("label", key),
                m.get("value"),
                m.get("formatted", ""),
                m.get("assessment", ""),
                m.get("benchmark", ""),
                m.get("detail", ""),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c, value=str(v) if v is not None else "")
                _style_data(cell, alt_row=alt, align="left")
            ws.row_dimensions[r_idx].height = 30
            current_row[0] += 1
        current_row[0] += 1   # blank spacer row

    def _list_block(items: list, col_headers: list, row_builder):
        if not items:
            return
        for c, h in enumerate(col_headers, 1):
            cell = ws.cell(row=current_row[0], column=c, value=h)
            _style_header(cell, level=3)
        ws.row_dimensions[current_row[0]].height = 16
        current_row[0] += 1
        for idx, item in enumerate(items):
            alt   = (idx % 2 == 1)
            r_idx = current_row[0]
            row_vals = row_builder(item)
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r_idx, column=c, value=str(v) if v is not None else "")
                sev   = item.get("severity") if isinstance(item, dict) else None
                if sev:
                    cell.fill = _fill(_SEVERITY_COLORS.get(sev, _C["white"]))
                    cell.font = _font(size=10)
                    cell.border = _border()
                    cell.alignment = _left()
                else:
                    _style_data(cell, alt_row=alt)
            ws.row_dimensions[r_idx].height = 30
            current_row[0] += 1
        current_row[0] += 1

    # ── Classification ────────────────────────────────────────────────────
    clf_result = r.get("classification_result") or {}
    _section("CLASSIFICATION")
    clf_row = current_row[0]
    ws.cell(row=clf_row, column=1, value="Verdict")
    ws.cell(row=clf_row, column=2, value=clf)
    clf_color = _CLASSIFICATION_COLORS.get(clf)
    if clf_color:
        for c in range(1, 7):
            ws.cell(row=clf_row, column=c).fill = _fill(clf_color)
            ws.cell(row=clf_row, column=c).font = _font(
                bold=True, color=_C["white"] if clf in ("STRONG BUY", "BUY", "AVOID", "STRONG AVOID") else _C["black"],
                size=11)
    ws.row_dimensions[clf_row].height = 20
    current_row[0] += 1

    for label, items in [("Positives", clf_result.get("reasons_for") or []),
                         ("Negatives", clf_result.get("reasons_against") or [])]:
        for i, reason in enumerate(items):
            cell1 = ws.cell(row=current_row[0], column=1, value=label if i == 0 else "")
            cell2 = ws.cell(row=current_row[0], column=2, value=reason)
            ws.merge_cells(f"B{current_row[0]}:F{current_row[0]}")
            alt = (i % 2 == 1)
            _style_data(cell1, alt_row=alt, bold=True)
            _style_data(cell2, alt_row=alt)
            ws.row_dimensions[current_row[0]].height = 24
            current_row[0] += 1
    current_row[0] += 1

    # ── Valuation ─────────────────────────────────────────────────────────
    _section("VALUATION")
    _metrics_block(r.get("valuation_metrics") or {})

    # ── Quality ───────────────────────────────────────────────────────────
    _section("BUSINESS QUALITY")
    _metrics_block(r.get("quality_metrics") or {})

    # ── Financial Strength ────────────────────────────────────────────────
    _section("FINANCIAL STRENGTH")
    _metrics_block(r.get("financial_metrics") or {})

    # ── Growth ────────────────────────────────────────────────────────────
    _section("GROWTH TRAJECTORY")
    _metrics_block(r.get("growth_metrics") or {})

    # ── Risk ──────────────────────────────────────────────────────────────
    _section("RISK PROFILE")
    _metrics_block(r.get("risk_metrics") or {})

    # ── Red Flags ─────────────────────────────────────────────────────────
    _section("RED FLAGS")
    _list_block(
        r.get("red_flags") or [],
        col_headers=["Severity", "Pattern", "Detail", "", "", ""],
        row_builder=lambda f: [f["severity"], f["pattern"], f["detail"], "", "", ""],
    )
    summary_cell = ws.cell(row=current_row[0], column=1,
                           value=r.get("red_flag_summary", ""))
    summary_cell.font      = _font(bold=True, italic=True, size=10)
    summary_cell.alignment = _left()
    ws.merge_cells(f"A{current_row[0]}:F{current_row[0]}")


# ── spec v2: new sheet builders ─────────────────────────────────────────────

def _safe_float(v) -> float | None:
    try:
        if v is None:
            return None
        return float(v)
    except Exception:
        return None


def _positive_float(v) -> float | None:
    """Return float only if v is strictly positive; otherwise treat as missing."""
    f = _safe_float(v)
    if f is None or f <= 0:
        return None
    return f


def _classification_dark(clf: str) -> bool:
    # White text for the darker backgrounds in our classification palette.
    return clf in ("STRONG BUY", "BUY", "AVOID", "STRONG AVOID")


def _timing_signal_from_record(record: dict) -> str:
    """
    Spec timing logic:
      - Oversold: RSI < 30 AND price < MA200
      - Uptrend: price > MA50 > MA200
      - Downtrend: MA50 < MA200
      - else: Neutral
    """
    price = _safe_float(record.get("current_price"))
    ma50 = _safe_float(record.get("ma50"))
    ma200 = _safe_float(record.get("ma200"))

    mom = record.get("momentum_metrics") or {}
    rsi_val = (mom.get("rsi") or {}).get("value")
    rsi = _safe_float(rsi_val)
    if rsi is None:
        rsi = _safe_float(record.get("rsi14"))

    if rsi is not None and price is not None and ma200 is not None:
        if rsi < 30 and price < ma200:
            return "Oversold"

    if price is not None and ma50 is not None and ma200 is not None:
        if price > ma50 and ma50 > ma200:
            return "Uptrend"
        if ma50 < ma200:
            return "Downtrend"

    return "Neutral"


def _style_sheet_title(ws, row: int, value: str, col_end: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=value)
    cell.font = _font(bold=True, size=14, color=_C["header_dark"])
    cell.alignment = _left()
    ws.row_dimensions[row].height = 22


def _style_table_header_cell(cell):
    _style_header(cell, level=1)
    cell.font = _font(bold=True, color=_C["white"], size=10)


def _write_cell(ws, row: int, col: int, value, *, alt: bool = False, bold: bool = False,
                align: str = "left", number_format: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    _style_data(cell, alt_row=alt, bold=bold, align=align)
    if number_format is not None and value is not None:
        cell.number_format = number_format
    return cell


def _build_dashboard(wb: "Workbook", results: list[dict]):
    ws = wb.create_sheet("Dashboard", 0)
    # Freeze header row + first two columns (Ticker, Company).
    ws.freeze_panes = "C3"
    _style_sheet_title(ws, 1, f"Stock Analyzer — {datetime.today().strftime('%Y-%m-%d')}", col_end=10)

    headers = [
        "Ticker", "Company", "Sector", "Price (native)", "Fair Value", "Upside %", "Classification",
        "Timing Signal", "Key Opportunity", "Key Risk",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        _style_table_header_cell(cell)
    widths = [10, 30, 20, 12, 12, 9, 14, 14, 45, 45]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    def _upside_for(r: dict) -> float | None:
        price = _positive_float(r.get("current_price"))
        fv = _positive_float(r.get("fair_value_weighted"))
        if price is None or fv is None:
            return None
        return (fv - price) / price

    sorted_results = sorted(results, key=lambda r: _upside_for(r) if _upside_for(r) is not None else -999, reverse=True)

    for i, r in enumerate(sorted_results):
        row = i + 3
        alt = (i % 2 == 1)

        ticker = r.get("ticker", "")
        company = (r.get("company_name") or "")[:40]
        sector = (r.get("sector") or "")[:30]
        price = _positive_float(r.get("current_price"))
        fv = _positive_float(r.get("fair_value_weighted"))
        upside = _upside_for(r)

        clf = (r.get("classification_result") or {}).get("classification", "N/A")
        timing_signal = _timing_signal_from_record(r)
        key_opp = (r.get("classification_result") or {}).get("key_opportunity", "")
        key_risk = (r.get("classification_result") or {}).get("key_risk", "")

        _write_cell(ws, row, 1, ticker, alt=alt, align="left")
        _write_cell(ws, row, 2, company, alt=alt, align="left")
        _write_cell(ws, row, 3, sector, alt=alt, align="left")
        _write_cell(ws, row, 4, round(price, 2) if price is not None else None, alt=alt, number_format="0.00")
        _write_cell(ws, row, 5, round(fv, 2) if fv is not None else None, alt=alt, number_format="0.00")
        _write_cell(ws, row, 6, upside if upside is not None else None, alt=alt, number_format="0.0%")

        # Classification cell styling.
        clf_cell = ws.cell(row=row, column=7, value=clf)
        _style_data(clf_cell, alt_row=alt, bold=False, align="center")
        clf_color = _CLASSIFICATION_COLORS.get(clf)
        if clf_color:
            clf_cell.fill = _fill(clf_color)
            clf_cell.font = _font(
                bold=True,
                color=_C["white"] if _classification_dark(clf) else _C["black"],
                size=10,
            )

        _write_cell(ws, row, 8, timing_signal, alt=alt, align="center")
        _write_cell(ws, row, 9, key_opp, alt=alt, align="left")
        _write_cell(ws, row, 10, key_risk, alt=alt, align="left")

        # Upside conditional coloring.
        if upside is not None:
            c = ws.cell(row=row, column=6)
            if upside > 0.15:
                c.fill = _fill(_C["buy"])
                c.font = _font(bold=True, color=_C["white"], size=10)
            elif upside < -0.15:
                c.fill = _fill(_C["avoid"])
                c.font = _font(bold=True, color=_C["white"], size=10)

        ws.row_dimensions[row].height = 22


def _build_scorecard(wb: "Workbook", results: list[dict]):
    ws = wb.create_sheet("Scorecard", 1)
    ws.freeze_panes = "C3"
    _style_sheet_title(ws, 1, "Scorecard — Numerical Comparison", col_end=17)

    headers = [
        "Ticker", "Company", "Currency", "Price",
        "ROIC", "WACC", "ROIC/WACC Spread",
        "Rev CAGR 5Y", "Op Margin", "Net Margin", "Gross Margin",
        "Debt/Equity", "ND/EBITDA", "Interest Coverage",
        "FCF Yield", "Beta", "Data Quality",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        _style_table_header_cell(cell)

    widths = [10, 30, 10, 12, 10, 10, 18, 12, 11, 11, 12, 12, 12, 16, 10, 8, 12]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    def _metric_fill(value, kind: str):
        if value is None:
            return None
        green = (_C["buy"], _C["white"])
        yellow = (_C["hold"], _C["black"])
        red = (_C["strong_avoid"], _C["white"])

        if kind == "roic":
            if value >= 0.15:
                return green
            if value >= 0.08:
                return yellow
            return red
        if kind == "spread":
            if value > 0.03:
                return green
            if value >= 0.0:
                return yellow
            return red
        if kind == "nd_ebitda":
            if value < 1.0:
                return green
            if value <= 3.0:
                return yellow
            return red
        if kind == "interest_coverage":
            if value > 8.0:
                return green
            if value >= 3.0:
                return yellow
            return red
        if kind == "fcf_yield":
            if value > 0.05:
                return green
            if value >= 0.02:
                return yellow
            return red
        return None

    for i, r in enumerate(results):
        row = i + 3
        alt = (i % 2 == 1)

        ticker = r.get("ticker", "")
        company = (r.get("company_name") or "")[:40]
        currency = r.get("currency", "")
        price = _safe_float(r.get("current_price"))

        roic = _safe_float(r.get("roic"))
        wacc = _safe_float((r.get("wacc_data") or {}).get("wacc"))
        spread = (roic - wacc) if (roic is not None and wacc is not None) else None

        rev_cagr = _safe_float(r.get("revenue_cagr_5y"))
        op_margin = _safe_float(r.get("operating_margin"))
        net_margin = _safe_float(r.get("net_margin"))
        gross_margin = _safe_float(r.get("gross_margin"))

        debt_equity = _safe_float(r.get("debt_equity"))
        nd_ebitda = _safe_float(r.get("net_debt_ebitda"))
        interest_coverage = _safe_float(r.get("interest_coverage"))
        fcf_yield = _safe_float(r.get("fcf_yield"))
        beta = _safe_float(r.get("beta"))
        dq = r.get("data_quality_score")
        dq = int(dq) if isinstance(dq, (int, float)) and dq is not None else None

        values = [
            (ticker, None, "left"),
            (company, None, "left"),
            (currency, None, "center"),
            (price, "0.00", "center"),
            (roic, "0.0%", "center"),
            (wacc, "0.0%", "center"),
            (spread, "0.0%", "center"),
            (rev_cagr, "0.0%", "center"),
            (op_margin, "0.0%", "center"),
            (net_margin, "0.0%", "center"),
            (gross_margin, "0.0%", "center"),
            (debt_equity, "0.0", "center"),
            (nd_ebitda, '0.0"x"', "center"),
            (interest_coverage, '0.0"x"', "center"),
            (fcf_yield, "0.0%", "center"),
            (beta, "0.00", "center"),
            (dq, "0", "center"),
        ]

        for c, (v, nf, align) in enumerate(values, 1):
            _write_cell(ws, row, c, v, alt=alt, align=align, number_format=nf)

        # Conditional fills per spec.
        checks = [
            (5, roic, "roic"),
            (7, spread, "spread"),
            (13, nd_ebitda, "nd_ebitda"),
            (14, interest_coverage, "interest_coverage"),
            (15, fcf_yield, "fcf_yield"),
        ]
        for col_idx, val, kind in checks:
            fill = _metric_fill(val, kind)
            if fill:
                fill_color, font_color = fill
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = _fill(fill_color)
                cell.font = _font(bold=True, color=font_color, size=10)

        ws.row_dimensions[row].height = 18


def _build_scenarios(wb: "Workbook", results: list[dict]):
    ws = wb.create_sheet("Scenarios", 2)
    _style_sheet_title(ws, 1, "Scenarios — Bear/Base/Bull DCF", col_end=4)
    # Freeze up through the first block's column headers (Row 4).
    ws.freeze_panes = "A5"

    _set_col_width(ws, 1, 50)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 16)

    row_ptr = 3
    base_case_tg = 0.025

    for r in results:
        ticker = r.get("ticker", "")
        company = r.get("company_name") or ticker
        scenarios = r.get("scenarios") or {}
        weights = scenarios.get("_weights") or {}
        bear = scenarios.get("bear") or {}
        base = scenarios.get("base") or {}
        bull = scenarios.get("bull") or {}

        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
        title_cell = ws.cell(row=row_ptr, column=1, value=f"{ticker} — {company}")
        _style_header(title_cell, level=1)
        title_cell.font = _font(bold=True, color=_C["white"], size=11)
        title_cell.alignment = _left()
        ws.row_dimensions[row_ptr].height = 22
        row_ptr += 1

        headers = ["Parameter", "Bear", "Base", "Bull"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row_ptr, column=c, value=h)
            _style_table_header_cell(cell)
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

        _write_cell(ws, row_ptr, 1, "Probability weight", bold=True, align="left")
        _write_cell(ws, row_ptr, 2, weights.get("bear"), bold=True, align="center", number_format="0.0%")
        _write_cell(ws, row_ptr, 3, weights.get("base"), bold=True, align="center", number_format="0.0%")
        _write_cell(ws, row_ptr, 4, weights.get("bull"), bold=True, align="center", number_format="0.0%")
        row_ptr += 1

        param_rows = [
            ("Starting revenue growth", "growth_rate", "0.0%"),
            ("Starting operating margin", "operating_margin", "0.0%"),
            ("Steady-state margin (yr 10)", "steady_state_margin", "0.0%"),
            ("Terminal growth rate", "terminal_growth", "0.0%"),
            ("WACC", "wacc", "0.0%"),
            ("PV of 10Y FCFs", "pv_fcfs", "0.00"),
            ("PV of terminal value", "pv_terminal_value", "0.00"),
            ("Terminal value % of EV", "tv_pct_of_ev", "0.0%"),
            ("Enterprise value", "enterprise_value", "0.00"),
            ("Equity value", "equity_value", "0.00"),
        ]

        money_keys = {"pv_fcfs", "pv_terminal_value", "enterprise_value", "equity_value"}
        for label, key, nf in param_rows:
            _write_cell(ws, row_ptr, 1, label, bold=False, align="left")
            if key in money_keys:
                _write_cell(ws, row_ptr, 2, _bn(bear.get(key)), align="center", number_format=None)
                _write_cell(ws, row_ptr, 3, _bn(base.get(key)), align="center", number_format=None)
                _write_cell(ws, row_ptr, 4, _bn(bull.get(key)), align="center", number_format=None)
            else:
                _write_cell(ws, row_ptr, 2, bear.get(key), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 3, base.get(key), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 4, bull.get(key), align="center", number_format=nf)
            ws.row_dimensions[row_ptr].height = 18
            row_ptr += 1

        _write_cell(ws, row_ptr, 1, "Fair value per share", bold=True, align="left")
        fv_b = _positive_float(bear.get("per_share_value"))
        fv_base = _positive_float(base.get("per_share_value"))
        fv_bull = _positive_float(bull.get("per_share_value"))
        _write_cell(ws, row_ptr, 2, fv_b, align="center", bold=True, number_format="0.00")
        _write_cell(ws, row_ptr, 3, fv_base, align="center", bold=True, number_format="0.00")
        _write_cell(ws, row_ptr, 4, fv_bull, align="center", bold=True, number_format="0.00")
        # Spec: Fair value per share should stand out (bold + larger font).
        for c in (1, 2, 3, 4):
            ws.cell(row=row_ptr, column=c).font = _font(bold=True, color=_C["black"], size=12)
        ws.row_dimensions[row_ptr].height = 20
        row_ptr += 1

        weighted_fv = _positive_float(r.get("fair_value_weighted"))
        price = _positive_float(r.get("current_price"))
        upside = ((weighted_fv - price) / price) if (weighted_fv is not None and price is not None) else None

        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
        _write_cell(ws, row_ptr, 1, "Probability-weighted fair value", bold=True, align="left")
        c = ws.cell(row=row_ptr, column=2, value=weighted_fv)
        if weighted_fv is not None:
            c.number_format = "0.00"
            c.fill = _fill(_C["medium_rf"])
            c.font = _font(bold=True, color=_C["black"], size=11)
            c.alignment = _center()
        ws.row_dimensions[row_ptr].height = 20
        row_ptr += 1

        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
        _write_cell(ws, row_ptr, 1, "Current price", bold=True, align="left")
        c = ws.cell(row=row_ptr, column=2, value=price)
        if price is not None:
            c.number_format = "0.00"
            c.alignment = _center()
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
        _write_cell(ws, row_ptr, 1, "Upside/downside vs weighted fair value", bold=True, align="left")
        c = ws.cell(row=row_ptr, column=2, value=upside)
        c.number_format = "0.0%"
        if upside is not None and upside > 0.15:
            c.fill = _fill(_C["buy"])
            c.font = _font(bold=True, color=_C["white"], size=10)
        elif upside is not None and upside < -0.15:
            c.fill = _fill(_C["avoid"])
            c.font = _font(bold=True, color=_C["white"], size=10)
        c.alignment = _center()
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

        # Sensitivity grid (if available).
        tv_sens = r.get("tv_sensitivity")
        if tv_sens and tv_sens.get("grid"):
            row_ptr += 2
            wacc_steps = tv_sens.get("wacc_steps") or []
            tg_steps = tv_sens.get("tg_steps") or []
            grid = tv_sens.get("grid") or []
            base_wacc = tv_sens.get("base_wacc")

            ws.cell(row=row_ptr, column=1, value="Sensitivity grid (per-share)")
            _style_table_header_cell(ws.cell(row=row_ptr, column=1))

            # tg headers
            for j, tg in enumerate(tg_steps, start=2):
                cell = ws.cell(row=row_ptr + 1, column=j, value=tg)
                cell.number_format = "0.0%"
                _style_table_header_cell(cell)

            # wacc headers
            for i, wc in enumerate(wacc_steps):
                cell = ws.cell(row=row_ptr + 2 + i, column=1, value=wc)
                cell.number_format = "0.0%"
                _style_table_header_cell(cell)

            highlight_i = None
            highlight_j = None
            if base_wacc is not None and wacc_steps and tg_steps:
                highlight_i = min(range(len(wacc_steps)), key=lambda idx: abs(wacc_steps[idx] - base_wacc))
                highlight_j = min(range(len(tg_steps)), key=lambda idx: abs(tg_steps[idx] - base_case_tg))

            for i in range(len(wacc_steps)):
                for j in range(len(tg_steps)):
                    val = None
                    try:
                        val = grid[i][j]
                    except Exception:
                        val = None
                    cell = ws.cell(row=row_ptr + 2 + i, column=2 + j, value=val)
                    if val is not None:
                        cell.number_format = "0.00"
                    cell.alignment = _center()
                    cell.border = _border()
                    if highlight_i == i and highlight_j == j:
                        cell.fill = _fill(_C["buy"])
                        cell.font = _font(bold=True, color=_C["white"], size=10)

            row_ptr = row_ptr + 2 + len(wacc_steps) + 1

        row_ptr += 2


def _build_peers(wb: "Workbook", peer_result: dict):
    ws = wb.create_sheet("Peer Comparison", 3)
    _style_sheet_title(ws, 1, "Peer Comparison — Sector Ranked Metrics", col_end=11)
    # Freeze up through the first block's column headers (Row 4).
    ws.freeze_panes = "A5"

    row_ptr = 3
    for group in peer_result.get("groups") or []:
        sector = group.get("sector") or "Unknown"
        peers = group.get("peers") or []

        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=11)
        hcell = ws.cell(row=row_ptr, column=1, value=sector)
        _style_header(hcell, level=1)
        hcell.font = _font(bold=True, color=_C["white"], size=11)
        hcell.alignment = _left()
        ws.row_dimensions[row_ptr].height = 22
        row_ptr += 1

        headers = [
            "Ticker", "Company", "Classification",
            "Rev CAGR 5Y", "Op Margin", "ROIC", "ND/EBITDA",
            "Int Coverage", "FCF Yield", "Fair Value", "Overall Score",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row_ptr, column=c, value=h)
            _style_table_header_cell(cell)
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

        for idx, p in enumerate(peers):
            alt = (idx % 2 == 1)
            ticker = p.get("ticker", "")
            company = p.get("company", "")
            clf = p.get("classification", "N/A")
            metrics = p.get("metrics") or {}
            overall = p.get("overall_percentile")

            _write_cell(ws, row_ptr, 1, ticker, alt=alt, align="left")
            _write_cell(ws, row_ptr, 2, company, alt=alt, align="left")

            clf_cell = ws.cell(row=row_ptr, column=3, value=clf)
            _style_data(clf_cell, alt_row=alt, align="center")
            clf_color = _CLASSIFICATION_COLORS.get(clf)
            if clf_color:
                clf_cell.fill = _fill(clf_color)
                clf_cell.font = _font(bold=True, color=_C["white"] if _classification_dark(clf) else _C["black"], size=10)

            # metric columns map to peer_engine metric keys
            col_map = [
                ("revenue_cagr_5y", 4),
                ("operating_margin", 5),
                ("roic", 6),
                ("net_debt_ebitda", 7),
                ("interest_coverage", 8),
                ("fcf_yield", 9),
                ("fair_value_weighted", 10),
            ]
            for mk, col_idx in col_map:
                m = metrics.get(mk) or {}
                formatted = m.get("formatted", "N/A")
                rank = m.get("rank")
                out_of = m.get("out_of")
                cell = ws.cell(row=row_ptr, column=col_idx, value=formatted)
                _style_data(cell, alt_row=alt, align="center")
                if rank == 1 and out_of and out_of > 1:
                    cell.font = _font(bold=True, color=_C["black"], size=10)

            score_cell = ws.cell(row=row_ptr, column=11, value=overall)
            _style_data(score_cell, alt_row=alt, align="center")
            score_cell.number_format = "0%"
            if overall is not None:
                if overall >= 0.70:
                    score_cell.fill = _fill(_C["buy"])
                    score_cell.font = _font(bold=True, color=_C["white"], size=10)
                elif overall >= 0.40:
                    score_cell.fill = _fill(_C["hold"])
                    score_cell.font = _font(bold=True, color=_C["black"], size=10)
                else:
                    score_cell.fill = _fill(_C["strong_avoid"])
                    score_cell.font = _font(bold=True, color=_C["white"], size=10)

            ws.row_dimensions[row_ptr].height = 18
            row_ptr += 1

        row_ptr += 2

    widths = [10, 30, 18, 14, 12, 10, 12, 15, 12, 16, 12]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)


def _build_technicals(wb: "Workbook", results: list[dict]):
    ws = wb.create_sheet("Technicals", 4)
    _style_sheet_title(ws, 1, "Technicals — Momentum & Timing Signals", col_end=16)
    # Freeze the header row (Row 2).
    ws.freeze_panes = "A3"

    headers = [
        "Ticker", "Price", "MA50", "MA200",
        "Price vs MA50", "Price vs MA200",
        "Trend", "RSI (14d)",
        "52W High", "52W Low", "% of 52W Range",
        "1M Return", "3M Return", "6M Return", "12M Return",
        "Timing Signal",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        _style_table_header_cell(cell)

    widths = [10, 10, 10, 10, 14, 14, 12, 10, 12, 12, 16, 10, 10, 10, 10, 14]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    for i, r in enumerate(results):
        row = i + 3
        alt = (i % 2 == 1)

        ticker = r.get("ticker", "")
        price = _safe_float(r.get("current_price"))
        ma50 = _safe_float(r.get("ma50"))
        ma200 = _safe_float(r.get("ma200"))
        price_vs_ma50 = ((price - ma50) / ma50) if (price is not None and ma50) else None
        price_vs_ma200 = ((price - ma200) / ma200) if (price is not None and ma200) else None

        mom = r.get("momentum_metrics") or {}
        trend = ((mom.get("moving_averages") or {}).get("formatted")) or "UNKNOWN"
        rsi = _safe_float((mom.get("rsi") or {}).get("value"))
        if rsi is None:
            rsi = _safe_float(r.get("rsi14"))

        w52_high = _safe_float(r.get("week52_high"))
        w52_low = _safe_float(r.get("week52_low"))
        range_pct = _safe_float((mom.get("range_52w") or {}).get("value"))

        ret_1m = _safe_float((mom.get("return_1m") or {}).get("value"))
        ret_3m = _safe_float((mom.get("return_3m") or {}).get("value"))
        ret_6m = _safe_float((mom.get("return_6m") or {}).get("value"))
        ret_12m = _safe_float((mom.get("return_12m") or {}).get("value"))

        timing_signal = _timing_signal_from_record(r)

        values = [
            (ticker, None, "left"),
            (price, "0.00", "center"),
            (ma50, "0.00", "center"),
            (ma200, "0.00", "center"),
            (price_vs_ma50, "0.0%", "center"),
            (price_vs_ma200, "0.0%", "center"),
            (trend, None, "center"),
            (int(round(rsi)) if rsi is not None else None, "0", "center"),
            (w52_high, "0.00", "center"),
            (w52_low, "0.00", "center"),
            (range_pct, "0.0%", "center"),
            (ret_1m, "0.0%", "center"),
            (ret_3m, "0.0%", "center"),
            (ret_6m, "0.0%", "center"),
            (ret_12m, "0.0%", "center"),
            (timing_signal, None, "center"),
        ]

        for c, (v, nf, align) in enumerate(values, 1):
            _write_cell(ws, row, c, v, alt=alt, align=align, number_format=nf)

        # Conditional formatting manual fills.
        if rsi is not None:
            cell = ws.cell(row=row, column=8)
            if rsi < 30:
                cell.fill = _fill(_C["buy"])
                cell.font = _font(bold=True, color=_C["white"], size=10)
            elif rsi > 70:
                cell.fill = _fill(_C["avoid"])
                cell.font = _font(bold=True, color=_C["white"], size=10)

        trend_cell = ws.cell(row=row, column=7)
        if isinstance(trend, str):
            t = trend.upper()
            if t == "UPTREND":
                trend_cell.fill = _fill(_C["buy"])
                trend_cell.font = _font(bold=True, color=_C["white"], size=10)
            elif t == "DOWNTREND":
                trend_cell.fill = _fill(_C["strong_avoid"])
                trend_cell.font = _font(bold=True, color=_C["white"], size=10)
            elif t == "SIDEWAYS":
                trend_cell.fill = _fill(_C["hold"])
                trend_cell.font = _font(bold=True, color=_C["black"], size=10)

        for col_idx, val in [(12, ret_1m), (13, ret_3m), (14, ret_6m), (15, ret_12m)]:
            if val is not None:
                cell = ws.cell(row=row, column=col_idx)
                if val >= 0:
                    cell.fill = _fill(_C["buy"])
                    cell.font = _font(bold=True, color=_C["white"], size=10)
                else:
                    cell.fill = _fill(_C["avoid"])
                    cell.font = _font(bold=True, color=_C["white"], size=10)

        ws.row_dimensions[row].height = 18


def _build_detail_sheet_v2(wb: "Workbook", record: dict, sheet_name: str):
    ws = wb.create_sheet(sheet_name)

    # Narrative readability.
    for c, w in zip(range(1, 7), [10, 30, 10, 10, 10, 10]):
        _set_col_width(ws, c, w)

    ws.merge_cells("A1:F1")
    one_liner = ((record.get("explanation") or {}).get("one_liner")) or ""
    c1 = ws.cell(row=1, column=1, value=one_liner)
    c1.font = _font(bold=False, italic=True, size=11, color=_C["header_dark"])
    c1.alignment = _left()
    ws.row_dimensions[1].height = 24

    row_ptr = 3

    def _section_header(title: str):
        nonlocal row_ptr
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        cell = ws.cell(row=row_ptr, column=1, value=title)
        _style_header(cell, level=2)
        cell.font = _font(bold=True, color=_C["white"], size=11)
        cell.alignment = _left()
        ws.row_dimensions[row_ptr].height = 22
        row_ptr += 1

    def _paragraph(heading: str, text: str):
        nonlocal row_ptr
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        hcell = ws.cell(row=row_ptr, column=1, value=heading)
        hcell.font = _font(bold=True, color=_C["header_dark"], size=10)
        hcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        tcell = ws.cell(row=row_ptr, column=1, value=text)
        tcell.font = _font(bold=False, color=_C["black"], size=10)
        tcell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        approx_lines = max(1, len(text) // 85)
        ws.row_dimensions[row_ptr].height = 16 + min(approx_lines, 5) * 10
        row_ptr += 1

    # A: Narrative summary.
    _section_header("Narrative Summary")
    paragraphs = ((record.get("explanation") or {}).get("paragraphs")) or []
    allowed = {"Business Overview", "Financial Health", "Valuation", "Key Risks", "Verdict"}
    for heading, text in paragraphs:
        if heading in allowed:
            _paragraph(heading, text)
    row_ptr += 1

    # B: Key numbers.
    _section_header("Key Numbers")
    wacc = _safe_float((record.get("wacc_data") or {}).get("wacc"))
    roic = _safe_float(record.get("roic"))
    spread = (roic - wacc) if (roic is not None and wacc is not None) else None

    fair_value = _positive_float(record.get("fair_value_weighted"))
    price = _positive_float(record.get("current_price"))
    buy_below = _positive_float(record.get("buy_below_price"))
    upside = ((fair_value - price) / price) if (fair_value is not None and price is not None) else None

    left_items = [
        ("Price", price, "0.00"),
        ("Fair Value", fair_value, "0.00"),
        ("Buy-Below", buy_below, "0.00"),
        ("Upside %", upside, "0.0%"),
        ("Classification", (record.get("classification_result") or {}).get("classification", "N/A"), None),
    ]
    right_items = [
        ("ROIC", roic, "0.0%"),
        ("WACC", wacc, "0.0%"),
        ("ROIC/WACC Spread", spread, "0.0%"),
        ("Rev CAGR 5Y", _safe_float(record.get("revenue_cagr_5y")), "0.0%"),
        ("Op Margin", _safe_float(record.get("operating_margin")), "0.0%"),
        ("ND/EBITDA", _safe_float(record.get("net_debt_ebitda")), '0.0"x"'),
    ]

    for idx in range(6):
        row = row_ptr + idx
        # left pair
        if idx < len(left_items):
            lbl, val, nf = left_items[idx]
            _write_cell(ws, row, 1, lbl, alt=False, bold=True, align="left")
            cell = ws.cell(row=row, column=2, value=val)
            _style_data(cell, alt_row=False, bold=False, align="left")
            if lbl == "Classification":
                clf = str(val)
                clf_color = _CLASSIFICATION_COLORS.get(clf)
                if clf_color:
                    cell.fill = _fill(clf_color)
                    cell.font = _font(bold=True, color=_C["white"] if _classification_dark(clf) else _C["black"], size=10)
            elif nf and val is not None:
                cell.number_format = nf
        # right pair
        if idx < len(right_items):
            lbl, val, nf = right_items[idx]
            _write_cell(ws, row, 3, lbl, alt=False, bold=True, align="left")
            cell = ws.cell(row=row, column=4, value=val)
            _style_data(cell, alt_row=False, bold=False, align="left")
            if nf and val is not None:
                cell.number_format = nf

        ws.row_dimensions[row].height = 18

    row_ptr += 6 + 1

    # C: Scenario summary.
    _section_header("Scenario Summary")
    scenarios = record.get("scenarios") or {}
    weights = scenarios.get("_weights") or {}
    bear = scenarios.get("bear") or {}
    base = scenarios.get("base") or {}
    bull = scenarios.get("bull") or {}

    headers = ["Parameter", "Bear", "Base", "Bull"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_ptr, column=c, value=h)
        _style_table_header_cell(cell)
    row_ptr += 1

    _write_cell(ws, row_ptr, 1, "Probability weight", bold=True, align="left")
    _write_cell(ws, row_ptr, 2, weights.get("bear"), bold=True, align="center", number_format="0.0%")
    _write_cell(ws, row_ptr, 3, weights.get("base"), bold=True, align="center", number_format="0.0%")
    _write_cell(ws, row_ptr, 4, weights.get("bull"), bold=True, align="center", number_format="0.0%")
    row_ptr += 1

    param_rows = [
        ("Starting revenue growth", "growth_rate", "0.0%"),
        ("Starting operating margin", "operating_margin", "0.0%"),
        ("Steady-state margin (yr 10)", "steady_state_margin", "0.0%"),
        ("Terminal growth rate", "terminal_growth", "0.0%"),
        ("WACC", "wacc", "0.0%"),
        ("PV of 10Y FCFs", "pv_fcfs", "0.00"),
        ("PV of terminal value", "pv_terminal_value", "0.00"),
        ("Terminal value % of EV", "tv_pct_of_ev", "0.0%"),
        ("Enterprise value", "enterprise_value", "0.00"),
        ("Equity value", "equity_value", "0.00"),
        ("Fair value per share", "per_share_value", "0.00"),
    ]

    money_keys = {"pv_fcfs", "pv_terminal_value", "enterprise_value", "equity_value"}
    for label, key, nf in param_rows:
        _write_cell(ws, row_ptr, 1, label, align="left")
        if key in money_keys:
            _write_cell(ws, row_ptr, 2, _bn(bear.get(key)), align="center", number_format=None)
            _write_cell(ws, row_ptr, 3, _bn(base.get(key)), align="center", number_format=None)
            _write_cell(ws, row_ptr, 4, _bn(bull.get(key)), align="center", number_format=None)
        else:
            # Avoid displaying negative intrinsic values as a "price".
            if key == "per_share_value":
                _write_cell(ws, row_ptr, 2, _positive_float(bear.get(key)), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 3, _positive_float(base.get(key)), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 4, _positive_float(bull.get(key)), align="center", number_format=nf)
            else:
                _write_cell(ws, row_ptr, 2, bear.get(key), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 3, base.get(key), align="center", number_format=nf)
                _write_cell(ws, row_ptr, 4, bull.get(key), align="center", number_format=nf)
        # Spec: Fair value per share should be bold and larger.
        if key == "per_share_value":
            for c in (1, 2, 3, 4):
                ws.cell(row=row_ptr, column=c).font = _font(bold=True, color=_C["black"], size=12)
        ws.row_dimensions[row_ptr].height = 18
        row_ptr += 1

    ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
    _write_cell(ws, row_ptr, 1, "Probability-weighted fair value", bold=True, align="left")
    c = ws.cell(row=row_ptr, column=2, value=fair_value)
    if fair_value is not None:
        c.number_format = "0.00"
        c.fill = _fill(_C["medium_rf"])
        c.font = _font(bold=True, color=_C["black"], size=11)
        c.alignment = _center()
    row_ptr += 1

    ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
    _write_cell(ws, row_ptr, 1, "Current price", bold=True, align="left")
    c = ws.cell(row=row_ptr, column=2, value=price)
    if price is not None:
        c.number_format = "0.00"
        c.alignment = _center()
    row_ptr += 1

    ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
    _write_cell(ws, row_ptr, 1, "Upside/downside vs weighted fair value", bold=True, align="left")
    c = ws.cell(row=row_ptr, column=2, value=upside)
    c.number_format = "0.0%"
    if upside is not None and upside > 0.15:
        c.fill = _fill(_C["buy"])
        c.font = _font(bold=True, color=_C["white"], size=10)
    elif upside is not None and upside < -0.15:
        c.fill = _fill(_C["avoid"])
        c.font = _font(bold=True, color=_C["white"], size=10)
    c.alignment = _center()
    row_ptr += 1

    row_ptr += 1

    # D: Red flags.
    red_flags = record.get("red_flags") or []
    if red_flags:
        _section_header("Red Flags")
        headers = ["Severity", "Pattern", "Detail"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row_ptr, column=c, value=h)
            _style_table_header_cell(cell)
        row_ptr += 1

        _set_col_width(ws, 1, 10)
        _set_col_width(ws, 2, 25)
        _set_col_width(ws, 3, 55)

        for f in red_flags:
            sev = f.get("severity", "LOW")
            pattern = f.get("pattern", "")
            detail = f.get("detail", "")

            _write_cell(ws, row_ptr, 1, sev, alt=False, align="center")
            sev_cell = ws.cell(row=row_ptr, column=1)
            sev_fill = _SEVERITY_COLORS.get(sev)
            if sev_fill:
                sev_cell.fill = _fill(sev_fill)
                sev_cell.font = _font(bold=True, color=_C["black"], size=10)

            _write_cell(ws, row_ptr, 2, pattern, alt=False, align="left")
            det = ws.cell(row=row_ptr, column=3, value=detail)
            _style_data(det, alt_row=False, align="left")
            det.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row_ptr].height = 20
            row_ptr += 1


# ── main function ──────────────────────────────────────────────────────────────

def write_excel(results: list[dict], filepath: str = None) -> str:
    """
    Build and save the workbook. Returns the filepath written.
    Raises ImportError if openpyxl is not installed.
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel output. "
            "Install it with:  pip install openpyxl"
        )

    if filepath is None:
        date_str = datetime.today().strftime("%Y%m%d_%H%M")
        filepath = f"stock_analysis_{date_str}.xlsx"

    wb = Workbook()
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_dashboard(wb, results)
    _build_scorecard(wb, results)
    _build_scenarios(wb, results)

    peer_result = analyze_peers(results)
    _build_peers(wb, peer_result)

    _build_technicals(wb, results)

    for r in results:
        ticker = r.get("ticker", "UNKNOWN")
        # Sheet names max 31 chars, no special chars
        safe_name = ticker.replace(".", "_").replace("/", "_").replace("\\", "_")
        _build_detail_sheet_v2(wb, r, safe_name[:31])

    wb.save(filepath)
    return filepath
